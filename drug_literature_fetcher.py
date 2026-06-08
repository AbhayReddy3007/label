"""
Drug Literature Fetcher — Multi-Source
Fetches scientific literature from PubMed, Europe PMC, Semantic Scholar,
OpenAlex, and CORE. Extracts disease indications via Gemini 2.5 Flash.

Output: one indication per row in standardized format (same columns as
label.py and drug_indication_researcher.py).

Usage:
    python drug_literature_fetcher.py --drug "Semaglutide"
    python drug_literature_fetcher.py --drug "Metformin" --sources pubmed,openalex
"""

import argparse
import os
import sys
import time
import json
import re
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

from indication_standardizer import standardize_indication, classify_indication, process_indications

# ── Load .env ─────────────────────────────────────────────────────────────────
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    env_file = Path(__file__).parent / ".env"
    if env_file.exists():
        for line in env_file.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, _, v = line.partition("=")
                os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

GEMINI_API_URL = (
    "https://generativelanguage.googleapis.com/v1beta/models/"
    "gemini-2.5-flash:generateContent"
)
OUTPUT_FILE = "journals.xlsx"
DEFAULT_MAX = 50

SOURCE_COLORS = {
    "PubMed":           "2E6DA4",
    "Europe PMC":       "217A4E",
    "Semantic Scholar": "8B4513",
    "OpenAlex":         "6A0DAD",
    "CORE":             "B8860B",
}


# ══════════════════════════════════════════════════════════════════════════════
#  FETCHERS (unchanged logic)
# ══════════════════════════════════════════════════════════════════════════════

def fetch_pubmed(drug: str, max_results: int = DEFAULT_MAX) -> list[dict]:
    SEARCH = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
    FETCH  = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
    BASE   = "https://pubmed.ncbi.nlm.nih.gov"
    try:
        r = requests.get(SEARCH, params={
            "db": "pubmed", "term": f'"{drug}"[Title/Abstract]',
            "retmax": max_results, "retmode": "json",
        }, timeout=15)
        r.raise_for_status()
        pmids = r.json()["esearchresult"]["idlist"]
        if not pmids:
            return []
        articles = []
        for i in range(0, len(pmids), 50):
            batch = pmids[i:i+50]
            resp = requests.get(FETCH, params={
                "db": "pubmed", "id": ",".join(batch),
                "retmode": "xml", "rettype": "abstract",
            }, timeout=30)
            resp.raise_for_status()
            articles.extend(_parse_pubmed_xml(resp.text, BASE))
            time.sleep(0.35)
        return articles
    except Exception as e:
        print(f"  ⚠️  PubMed error: {e}")
        return []


def _parse_pubmed_xml(xml_text: str, base_url: str) -> list[dict]:
    articles = []
    for block in re.findall(r"<PubmedArticle>(.*?)</PubmedArticle>", xml_text, re.DOTALL):
        pmid    = _rx(r"<PMID[^>]*>(.*?)</PMID>", block)
        title   = _strip_tags(_rx(r"<ArticleTitle>(.*?)</ArticleTitle>", block, re.DOTALL))
        journal = _strip_tags(_rx(r"<Title>(.*?)</Title>", block, re.DOTALL))
        abstract_parts = re.findall(r"<AbstractText[^>]*>(.*?)</AbstractText>", block, re.DOTALL)
        abstract = _strip_tags(" ".join(abstract_parts))
        if not abstract or not pmid:
            continue
        articles.append({
            "source": "PubMed", "title": title, "journal": journal,
            "abstract": abstract, "link": f"{base_url}/{pmid}/",
            "dedup_key": pmid,
        })
    return articles


def fetch_europepmc(drug: str, max_results: int = DEFAULT_MAX) -> list[dict]:
    BASE = "https://www.ebi.ac.uk/europepmc/webservices/rest/search"
    try:
        articles, page, page_size = [], 1, min(max_results, 100)
        while len(articles) < max_results:
            r = requests.get(BASE, params={
                "query": f'ABSTRACT:"{drug}" HAS_ABSTRACT:Y',
                "resultType": "core", "format": "json",
                "pageSize": page_size, "page": page,
            }, timeout=20)
            r.raise_for_status()
            results = r.json().get("resultList", {}).get("result", [])
            if not results:
                break
            for item in results:
                abstract = item.get("abstractText", "").strip()
                if not abstract:
                    continue
                pmid = item.get("pmid") or item.get("id", "")
                articles.append({
                    "source": "Europe PMC",
                    "title":   item.get("title", "N/A"),
                    "journal": item.get("journalTitle", "N/A"),
                    "abstract": abstract,
                    "link": f"https://europepmc.org/article/{item.get('source','MED')}/{item.get('id','')}",
                    "dedup_key": pmid or item.get("id", ""),
                })
            if len(results) < page_size:
                break
            page += 1
            time.sleep(0.3)
        return articles[:max_results]
    except Exception as e:
        print(f"  ⚠️  Europe PMC error: {e}")
        return []


def fetch_semantic_scholar(drug: str, max_results: int = DEFAULT_MAX) -> list[dict]:
    BASE = "https://api.semanticscholar.org/graph/v1/paper/search"
    try:
        articles, offset, limit = [], 0, min(max_results, 100)
        while len(articles) < max_results:
            r = requests.get(BASE, params={
                "query": drug, "limit": limit, "offset": offset,
                "fields": "title,abstract,journal,externalIds,url",
            }, timeout=20)
            if r.status_code == 429:
                time.sleep(5)
                continue
            r.raise_for_status()
            papers = r.json().get("data", [])
            if not papers:
                break
            for p in papers:
                abstract = (p.get("abstract") or "").strip()
                if not abstract:
                    continue
                pmid  = (p.get("externalIds") or {}).get("PubMed", "")
                ss_id = p.get("paperId", "")
                articles.append({
                    "source": "Semantic Scholar",
                    "title":   p.get("title", "N/A"),
                    "journal": (p.get("journal") or {}).get("name", "N/A"),
                    "abstract": abstract,
                    "link":   p.get("url") or f"https://www.semanticscholar.org/paper/{ss_id}",
                    "dedup_key": pmid or ss_id,
                })
            if len(papers) < limit:
                break
            offset += limit
            time.sleep(0.5)
        return articles[:max_results]
    except Exception as e:
        print(f"  ⚠️  Semantic Scholar error: {e}")
        return []


def fetch_openalex(drug: str, max_results: int = DEFAULT_MAX) -> list[dict]:
    BASE = "https://api.openalex.org/works"
    try:
        articles, page, per_page = [], 1, min(max_results, 50)
        while len(articles) < max_results:
            r = requests.get(BASE, params={
                "search": drug, "filter": "has_abstract:true",
                "per-page": per_page, "page": page,
                "select": "id,title,abstract_inverted_index,primary_location,doi",
            }, headers={"User-Agent": "DrugLitFetcher/2.0 (research tool)"}, timeout=20)
            r.raise_for_status()
            works = r.json().get("results", [])
            if not works:
                break
            for w in works:
                abstract = _reconstruct_openalex_abstract(w.get("abstract_inverted_index", {}))
                if not abstract:
                    continue
                loc    = w.get("primary_location") or {}
                source = loc.get("source") or {}
                doi    = w.get("doi") or ""
                link   = f"https://doi.org/{doi.replace('https://doi.org/', '')}" if doi else w.get("id", "")
                articles.append({
                    "source": "OpenAlex",
                    "title":   w.get("title", "N/A"),
                    "journal": source.get("display_name", "N/A"),
                    "abstract": abstract,
                    "link":    link,
                    "dedup_key": doi or w.get("id", ""),
                })
            if len(works) < per_page:
                break
            page += 1
            time.sleep(0.3)
        return articles[:max_results]
    except Exception as e:
        print(f"  ⚠️  OpenAlex error: {e}")
        return []


def _reconstruct_openalex_abstract(inverted_index: dict) -> str:
    if not inverted_index:
        return ""
    pos_word = {}
    for word, positions in inverted_index.items():
        for pos in positions:
            pos_word[pos] = word
    return " ".join(pos_word[p] for p in sorted(pos_word))


def fetch_core(drug: str, max_results: int = DEFAULT_MAX) -> list[dict]:
    BASE = "https://api.core.ac.uk/v3/search/works"
    try:
        r = requests.post(BASE, json={
            "q": drug, "limit": min(max_results, 100),
            "fields": ["title", "abstract", "journals", "sourceFulltextUrls", "doi"],
        }, timeout=20)
        r.raise_for_status()
        articles = []
        for item in r.json().get("results", []):
            abstract = (item.get("abstract") or "").strip()
            if not abstract:
                continue
            doi  = item.get("doi") or ""
            urls = item.get("sourceFulltextUrls") or []
            link = f"https://doi.org/{doi}" if doi else (urls[0] if urls else "https://core.ac.uk")
            journals     = item.get("journals") or []
            journal_name = journals[0].get("title", "N/A") if journals else "N/A"
            articles.append({
                "source": "CORE",
                "title":   item.get("title", "N/A"),
                "journal": journal_name,
                "abstract": abstract,
                "link":    link,
                "dedup_key": doi or str(item.get("id", "")),
            })
        return articles[:max_results]
    except Exception as e:
        print(f"  ⚠️  CORE error: {e}")
        return []


# ══════════════════════════════════════════════════════════════════════════════
#  Deduplication & Utilities
# ══════════════════════════════════════════════════════════════════════════════

def deduplicate(articles: list[dict]) -> list[dict]:
    seen_keys, seen_titles, unique = set(), [], []
    for a in articles:
        key = (a.get("dedup_key") or "").strip().lower()
        if key and key in seen_keys:
            continue
        if key:
            seen_keys.add(key)
        title_norm = re.sub(r"\W+", " ", (a.get("title") or "")).lower().strip()
        if any(_jaccard(title_norm, t) > 0.85 for t in seen_titles):
            continue
        seen_titles.append(title_norm)
        unique.append(a)
    return unique


def _jaccard(a: str, b: str) -> float:
    sa, sb = set(a.split()), set(b.split())
    if not sa or not sb:
        return 0.0
    return len(sa & sb) / len(sa | sb)


def _rx(pattern, text, flags=0):
    m = re.search(pattern, text, flags)
    return m.group(1).strip() if m else ""


def _strip_tags(text):
    return re.sub(r"<[^>]+>", "", text).strip()


# ══════════════════════════════════════════════════════════════════════════════
#  Gemini Indication Extraction
# ══════════════════════════════════════════════════════════════════════════════

def extract_indications(drug: str, title: str, abstract: str, api_key: str) -> list[str]:
    prompt = (
        f"You are a biomedical expert.\n\n"
        f"Drug: {drug}\nPaper title: {title}\nAbstract:\n{abstract}\n\n"
        f"Task: List ONLY the diseases, medical conditions, or clinical indications "
        f"this abstract mentions in the context of treating, preventing, or studying with {drug}.\n\n"
        f"Rules:\n"
        f"- Output a JSON array of short strings, e.g. [\"Type 2 Diabetes\",\"Obesity\"]\n"
        f"- Each entry: 1-5 words, title case\n"
        f"- Only include conditions where {drug} is a treatment/therapy/study subject\n"
        f"- If none found, return []\n"
        f"- Output ONLY valid JSON — no markdown, no explanation"
    )
    try:
        r = requests.post(
            f"{GEMINI_API_URL}?key={api_key}",
            headers={"Content-Type": "application/json"},
            json={
                "contents": [{"parts": [{"text": prompt}]}],
                "generationConfig": {"temperature": 0.1, "maxOutputTokens": 512},
            },
            timeout=30,
        )
        r.raise_for_status()
        raw = (
            r.json().get("candidates", [{}])[0]
             .get("content", {}).get("parts", [{}])[0]
             .get("text", "[]").strip()
        )
        raw = re.sub(r"^```[a-z]*\n?", "", raw)
        raw = re.sub(r"\n?```$", "", raw)
        parsed = json.loads(raw)
        if isinstance(parsed, list):
            return [str(x).strip() for x in parsed if str(x).strip()]
    except Exception:
        pass
    return []


# ══════════════════════════════════════════════════════════════════════════════
#  Source registry
# ══════════════════════════════════════════════════════════════════════════════

SOURCES = {
    "pubmed":          ("PubMed",           fetch_pubmed),
    "europepmc":       ("Europe PMC",       fetch_europepmc),
    "semanticscholar": ("Semantic Scholar", fetch_semantic_scholar),
    "openalex":        ("OpenAlex",         fetch_openalex),
    "core":            ("CORE",             fetch_core),
}


# ══════════════════════════════════════════════════════════════════════════════
#  COMMON FORMAT HEADERS (same as label.py / drug_indication_researcher.py)
# ══════════════════════════════════════════════════════════════════════════════

HEADERS    = ["molecule_name", "company_name", "indication", "indication_type",
              "trial_title", "trial_id", "phase", "source_url", "data_source"]
COL_WIDTHS = [18, 22, 28, 16, 50, 18, 10, 40, 16]


def _thin_border():
    return Border(bottom=Side(style="thin", color="DDDDDD"),
                  right=Side(style="thin", color="DDDDDD"))


# ══════════════════════════════════════════════════════════════════════════════
#  Excel Writer — common format
# ══════════════════════════════════════════════════════════════════════════════

def write_excel(rows: list[dict], drug: str, source_stats: dict) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Literature"

    thin     = _thin_border()
    hfill    = PatternFill("solid", start_color="1A1A2E")
    alt_fill = PatternFill("solid", start_color="F4F7FB")
    ncols    = len(HEADERS)

    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value     = f"Drug Literature  —  {drug.title()}"
    c.font      = Font(name="Arial", bold=True, size=14, color="1A1A2E")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    for col, h in enumerate(HEADERS, 1):
        c           = ws.cell(row=2, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = hfill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin
    ws.row_dimensions[2].height = 22

    for idx, row in enumerate(rows, start=3):
        fill = alt_fill if idx % 2 == 1 else None
        for col, key in enumerate(HEADERS, 1):
            c           = ws.cell(row=idx, column=col, value=row.get(key, ""))
            c.font      = Font(name="Arial", size=9)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border    = thin
            if fill:
                c.fill = fill
        ws.row_dimensions[idx].height = 18

    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}{2 + len(rows)}"
    ws.freeze_panes = "A3"
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUTPUT_FILE)
    print(f"\n✅  Saved → {OUTPUT_FILE}  ({len(rows)} rows)\n")


# ══════════════════════════════════════════════════════════════════════════════
#  Main
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Fetch drug literature from multiple sources; extract indications via Gemini 2.5 Flash."
    )
    parser.add_argument("--drug",    required=True, help="Drug name, e.g. Semaglutide")
    parser.add_argument("--sources", default="all",
                        help=f"Comma-separated or 'all'. Options: {', '.join(SOURCES)}")
    args = parser.parse_args()

    api_key = os.getenv("GEMINI_API_KEY", "").strip()
    if not api_key:
        sys.exit("❌  GEMINI_API_KEY not found in .env")

    drug = args.drug.strip()

    if args.sources.lower() == "all":
        active = list(SOURCES.keys())
    else:
        active = [s.strip().lower() for s in args.sources.split(",") if s.strip().lower() in SOURCES]
        if not active:
            sys.exit(f"❌  No valid sources. Choose from: {', '.join(SOURCES)}")

    print(f"\n{'━'*62}")
    print(f"  Drug    : {drug.title()}")
    print(f"  Sources : {', '.join(SOURCES[s][0] for s in active)}")
    print(f"  Output  : {OUTPUT_FILE}")
    print(f"{'━'*62}\n")

    # ── Fetch ─────────────────────────────────────────────────────────────────
    all_articles, source_stats = [], {}
    for key in active:
        label, fn = SOURCES[key]
        print(f"📡  {label} …")
        fetched = fn(drug)
        for a in fetched:
            a["source"] = label
        source_stats[label] = len(fetched)
        all_articles.extend(fetched)
        print(f"    → {len(fetched)} articles")

    print(f"\n🔗  Total raw        : {len(all_articles)}")
    all_articles = deduplicate(all_articles)
    print(f"✂️   After dedup      : {len(all_articles)}\n")

    # ── Gemini → standardize → unnest ─────────────────────────────────────────
    print("🤖  Extracting indications with Gemini 2.5 Flash …\n")
    rows = []
    for i, article in enumerate(all_articles, 1):
        print(f"  [{i:>3}/{len(all_articles)}] [{article['source']}] {article['title'][:65]}")
        raw_inds = extract_indications(drug, article["title"], article["abstract"], api_key)
        time.sleep(0.2)

        processed = process_indications(drug, raw_inds)
        if not processed:
            processed = [{"indication": "No indication found", "indication_type": ""}]

        for ind in processed:
            rows.append({
                "molecule_name":   drug.title(),
                "company_name":    article["source"],   # literature source as "company"
                "indication":      ind["indication"],
                "indication_type": ind["indication_type"],
                "trial_title":     article["title"],
                "trial_id":        "",                   # no trial_id for literature
                "phase":           "",                   # no phase for literature
                "source_url":      article["link"],
                "data_source":     "Literature",
            })

    # ── Excel ─────────────────────────────────────────────────────────────────
    print(f"\n📊  Writing {OUTPUT_FILE} …")
    write_excel(rows, drug, source_stats)


if __name__ == "__main__":
    main()
