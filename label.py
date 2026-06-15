"""
Clinical Efficacy — label.py
Fetches clinical trial data from BigQuery, enriches with Gemini,
then uses a SEPARATE Gemini + Google Search call to classify
indications as Primary / Secondary and assign therapy areas.

Output: one indication per row with Ep and Et scores.

Usage:
    python label.py semaglutide
"""

import sys
import os
import json
import re
import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from dotenv import load_dotenv
from google.cloud import bigquery
from google.oauth2 import service_account
from google import genai
from google.genai import types
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from indication_standardizer import standardize_indication

load_dotenv(override=True)

GEMINI_API_KEY  = os.getenv("GEMINI_API_KEY")
GBQ_PROJECT     = os.getenv("GBQ_PROJECT")
GBQ_DATASET     = os.getenv("GBQ_DATASET")
GBQ_TABLE       = os.getenv("GBQ_TABLE", "clinical_efficacy")
GBQ_SERVICE_KEY = os.getenv(
    "GBQ_SERVICE_KEY",
    r"C:\Users\p90022569\Downloads\cognito-prod-394707-d38a0283cb16 (2).json",
)

gemini_client = genai.Client(api_key=GEMINI_API_KEY)


def _safe_response_text(resp) -> str:
    """Safely extract text from a Gemini response.

    With Google Search grounding the simple `resp.text` property can be
    None or raise an error.  Walk the full candidates → parts tree and
    concatenate every text fragment we find.
    """
    # ── Fast path: .text works ────────────────────────────────────────────
    try:
        if resp.text is not None:
            return resp.text.strip()
    except Exception:
        pass

    # ── Slow path: collect text from all parts ────────────────────────────
    texts = []
    try:
        for candidate in (resp.candidates or []):
            try:
                parts = candidate.content.parts
            except Exception:
                continue
            for part in (parts or []):
                try:
                    t = getattr(part, "text", None)
                    if t:
                        texts.append(t.strip())
                except Exception:
                    continue
    except Exception:
        pass

    if texts:
        return "\n".join(texts)

    return ""


def _gemini_generate(client, gen_types, prompt, *, system_instruction="",
                     use_search=True, model="gemini-2.5-flash"):
    """Call Gemini with optional Google Search grounding.
    Retries on transient errors (503, 429, connection) with exponential backoff.
    If use_search is True and the response is empty, retries without grounding."""
    import time

    MAX_RETRIES = 4
    BASE_DELAY  = 5  # seconds

    configs = []
    if use_search:
        configs.append(gen_types.GenerateContentConfig(
            temperature=0,
            tools=[gen_types.Tool(google_search=gen_types.GoogleSearch())],
            system_instruction=system_instruction or "Return ONLY valid JSON.",
        ))
    configs.append(gen_types.GenerateContentConfig(
        temperature=0,
        system_instruction=system_instruction or "Return ONLY valid JSON.",
    ))

    def _is_transient(exc):
        err_str = str(exc).lower()
        return any(k in err_str for k in ("503", "429", "unavailable", "overloaded",
                                           "resource exhausted", "rate limit",
                                           "deadline exceeded", "connection",
                                           "timeout", "502", "500"))

    for i, cfg in enumerate(configs):
        last_err = None
        for attempt in range(MAX_RETRIES):
            try:
                resp = client.models.generate_content(
                    model=model, contents=prompt, config=cfg,
                )
                text = _safe_response_text(resp)
                text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text).strip()
                if text:
                    return text
                # Empty response — break out of retry loop to try next config
                break
            except Exception as e:
                last_err = e
                if _is_transient(e) and attempt < MAX_RETRIES - 1:
                    delay = BASE_DELAY * (2 ** attempt)
                    print(f"      ⏳ {e} — retrying in {delay}s ({attempt+1}/{MAX_RETRIES})…")
                    time.sleep(delay)
                elif i == 0 and len(configs) > 1:
                    print(f"      ↻  Error with Search ({e}) — trying without grounding…")
                    break  # break retry loop, move to next config
                else:
                    raise
        else:
            # All retries exhausted for this config
            if i == 0 and len(configs) > 1:
                print(f"      ↻  Retries exhausted — trying without grounding…")
                continue
            if last_err:
                raise last_err

        if i == 0 and len(configs) > 1:
            print("      ↻  Empty response with Search — retrying without grounding…")

    return ""


def _extract_json(text: str) -> dict | list:
    """Extract and parse the first JSON object or array from *text*,
    even if surrounded by prose, markdown fences, or whitespace."""
    if not text:
        raise ValueError("Cannot extract JSON from empty text")

    # 1. Direct parse
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # 2. Strip markdown fences
    stripped = re.sub(r"^```(?:json)?\s*", "", text, flags=re.MULTILINE)
    stripped = re.sub(r"\s*```\s*$", "", stripped, flags=re.MULTILINE).strip()
    try:
        return json.loads(stripped)
    except json.JSONDecodeError:
        pass

    # 3. Bracket-match the first { … } or [ … ]
    for open_ch, close_ch in (("{", "}"), ("[", "]")):
        start = stripped.find(open_ch)
        if start == -1:
            continue
        depth, end = 0, start
        in_str = False
        while end < len(stripped):
            ch = stripped[end]
            if ch == '"' and (end == 0 or stripped[end - 1] != '\\'):
                in_str = not in_str
            elif not in_str:
                if ch == open_ch:
                    depth += 1
                elif ch == close_ch:
                    depth -= 1
                    if depth == 0:
                        try:
                            return json.loads(stripped[start:end + 1])
                        except json.JSONDecodeError:
                            break
            end += 1

    raise ValueError(f"No valid JSON found in response (first 200 chars): {text[:200]}")


# ══════════════════════════════════════════════════════════════════════════════
#  PARALLEL CONFIG
# ══════════════════════════════════════════════════════════════════════════════

MAX_WORKERS = 10


# ══════════════════════════════════════════════════════════════════════════════
#  SECONDARY INDICATION CRITERIA
# ══════════════════════════════════════════════════════════════════════════════

SECONDARY_INDICATION_CRITERIA = """
A secondary indication qualifies ONLY if ALL of the following are true:

- The indication represents a true expansion — i.e., it is not part of the
  primary indication (for clinical assets) or currently approved label
  (for commercial assets)
- The indication is described at a clear disease-level definition, avoiding
  vague, overlapping, or synonymous representations
- The source must describe observed or measured outcomes in that specific
  indication (e.g., trial results, endpoint readouts, biomarker response),
  not just planned evaluation or exploratory intent

The following must NOT be considered secondary indications:
* Indications mentioned only as hypothesis, targets, or exploratory possibilities
* Pipeline indications without any data or outcomes
* Mechanism based assumptions without clinical or empirical validation
* Early discovery or preclinical signals without human data
* Any indication lacking traceable, verifiable evidence of results
"""


def _bq_client():
    credentials = service_account.Credentials.from_service_account_file(
        GBQ_SERVICE_KEY,
        scopes=["https://www.googleapis.com/auth/cloud-platform"],
    )
    return bigquery.Client(project=GBQ_PROJECT, credentials=credentials)

def load_checkpoint(file_path):
    if os.path.exists(file_path):
        with open(file_path, "r") as f:
            return json.load(f)
    return {}

def save_checkpoint(file_path, data):
    with open(file_path, "w") as f:
        json.dump(data, f, indent=2)

def fetch_rows(molecule_name):
    print("🔹 Connecting to BigQuery...")
    client = _bq_client()
    query = f"""
        SELECT molecule_name, company_name, source_url, phase, trial_id
        FROM `{GBQ_PROJECT}.{GBQ_DATASET}.{GBQ_TABLE}`
        WHERE LOWER(molecule_name) = LOWER(@molecule)
    """
    job_config = bigquery.QueryJobConfig(
        query_parameters=[bigquery.ScalarQueryParameter("molecule", "STRING", molecule_name)]
    )
    print("🔹 Running query...")
    results = client.query(query, job_config=job_config).result()
    rows = [dict(row) for row in results]
    print(f"✅ Retrieved {len(rows)} rows\n")
    return rows


# ══════════════════════════════════════════════════════════════════════════════
#  LLM CLASSIFICATION — one call per drug, with Google Search grounding
# ══════════════════════════════════════════════════════════════════════════════

def classify_indications_with_llm(molecule_name, unique_indications):
    """
    Uses Gemini + Google Search to research the drug and classify each
    indication as Primary / Secondary, and assign therapy areas.

    Returns dict:  indication_name → {"indication_type": ..., "therapy_area": ...}
    """
    if not unique_indications:
        return {}

    indications_json = json.dumps(unique_indications, indent=2)

    prompt = f"""
You are a pharmaceutical analyst. Your task is to research the drug
"{molecule_name}" and classify each of the following indications.

Drug / Molecule: {molecule_name}

Indications to classify:
{indications_json}

═══ STEP 1 — Research the drug ═══
Search the web to find:
  • What "{molecule_name}" is primarily approved / developed for
  • The FDA / EMA approved labels for this drug
  • The originator company's pipeline and clinical development programs

═══ STEP 2 — Classify each indication ═══
For EACH indication in the list above:

  indication_type:
    "Primary"   — this IS one of the drug's main approved or originally
                   intended indications (what it was designed and first
                   developed / approved to treat).
    "Secondary" — this is a label expansion or additional indication
                   beyond the primary use.
                   {SECONDARY_INDICATION_CRITERIA}

  therapy_area:
    Look at the INDICATION itself (the disease / condition) and determine
    which medical specialty treats it.
    Choose from: Metabolic, Cardiovascular, Oncology, Neuroscience,
    Immunology, Respiratory, Nephrology, Hepatology, Ophthalmology,
    Musculoskeletal, Gastroenterology, Infectious Disease, Dermatology,
    Hematology, Endocrinology, Rare Disease, or another appropriate
    broad therapeutic area.

    Examples:
      T2DM / Type 2 Diabetes / Obesity → Metabolic
      Heart Failure / MACE / CV Risk Reduction → Cardiovascular
      CKD / Diabetic Kidney Disease → Nephrology
      NASH / MASH → Hepatology
      NSCLC / Breast Cancer → Oncology
      Alzheimer's Disease → Neuroscience
      Rheumatoid Arthritis / Psoriasis → Immunology
      COPD / Asthma → Respiratory
      OSA (Obstructive Sleep Apnea) → Respiratory

Return ONLY valid JSON — no markdown fences, no explanation:
{{
  "classifications": [
    {{
      "indication": "<exact indication name from input list>",
      "indication_type": "Primary" or "Secondary",
      "therapy_area": "<therapy area>",
      "rationale": "<explain WHY this indication is Primary or Secondary for this drug — cite the specific evidence you found, e.g. FDA approval, original label, known label expansion, clinical data>"
    }}
  ]
}}
"""

    print(f"🔹 Classifying {len(unique_indications)} unique indications for {molecule_name} (LLM + Search)…")
    try:
        text = _gemini_generate(
            gemini_client, types, prompt,
            system_instruction=(
                "You are a pharmaceutical analyst. "
                "Search the web to find what this drug is approved for. "
                "Return ONLY valid JSON — no markdown, no explanation."
            ),
            use_search=True,
        )
        data = _extract_json(text)
        classifications = data.get("classifications", [])

        result = {}
        for c in classifications:
            name = c.get("indication", "").strip()
            if name:
                result[name.lower()] = {
                    "indication_type": c.get("indication_type", "Secondary"),
                    "therapy_area":    c.get("therapy_area", "Other"),
                    "rationale":       c.get("rationale", ""),
                }
                print(f"   • {name}: {c.get('indication_type')} | {c.get('therapy_area')}")

        # Retry missing indications individually
        for ind in unique_indications:
            if ind.lower() not in result:
                print(f"   🔎 Missing '{ind}' — retrying individually…")
                result[ind.lower()] = _classify_single_indication(
                    molecule_name, ind, gemini_client, types
                )

        return result

    except Exception as e:
        print(f"   ❌ Bulk classification failed: {e}")
        print(f"   🔎 Classifying each indication individually…")
        result = {}
        for ind in unique_indications:
            result[ind.lower()] = _classify_single_indication(
                molecule_name, ind, gemini_client, types
            )
        return result


def _classify_single_indication(molecule_name, indication, client, gen_types):
    """Classify a single indication via LLM + Google Search."""
    prompt = f"""
You are a pharmaceutical analyst. Research the drug "{molecule_name}".

Drug: {molecule_name}
Indication: {indication}

1. Is this indication "Primary" or "Secondary" for this drug?
   - "Primary": one of the drug's main approved / originally intended indications
   - "Secondary": a label expansion beyond the primary use

2. What therapy area does this INDICATION belong to?
   Choose from: Metabolic, Cardiovascular, Oncology, Neuroscience, Immunology,
   Respiratory, Nephrology, Hepatology, Ophthalmology, Musculoskeletal,
   Gastroenterology, Infectious Disease, Dermatology, Hematology,
   Endocrinology, Rare Disease, or another appropriate area.

3. Explain your reasoning.

Return ONLY valid JSON:
{{"indication_type": "Primary" or "Secondary", "therapy_area": "<area>", "rationale": "<why>"}}
"""
    try:
        text = _gemini_generate(
            client, gen_types, prompt,
            system_instruction="Return ONLY valid JSON.",
            use_search=True,
        )
        data = _extract_json(text)
        cls = {
            "indication_type": data.get("indication_type", "Secondary"),
            "therapy_area":    data.get("therapy_area", "Other"),
            "rationale":       data.get("rationale", ""),
        }
        print(f"     • {indication}: {cls['indication_type']} | {cls['therapy_area']}")
        return cls
    except Exception as ex:
        print(f"     ❌ Individual classification failed for '{indication}': {ex}")
        return {"indication_type": "Secondary", "therapy_area": "Other",
                "rationale": f"Classification failed: {ex}"}


# ══════════════════════════════════════════════════════════════════════════════
#  Ep SCORE — row-wise, based on phase
# ══════════════════════════════════════════════════════════════════════════════

def compute_ep(phase) -> str:
    """Phase 4 or 3 → 5, Phase 2 → 4, Phase 1 → 3."""
    phase_str = str(phase).strip().lower() if phase else ""
    nums = re.findall(r'\d+', phase_str)
    if not nums:
        if any(kw in phase_str for kw in ("approved", "launched", "marketed")):
            return "5"
        return ""
    max_phase = max(int(n) for n in nums)
    if max_phase >= 3:
        return "5"
    elif max_phase == 2:
        return "4"
    elif max_phase == 1:
        return "3"
    return ""


# ══════════════════════════════════════════════════════════════════════════════
#  Et SCORE — drug-level
# ══════════════════════════════════════════════════════════════════════════════

def compute_et(enriched_rows, molecule_name):
    """
    Multiple therapy areas                              → 5
    1 therapy area, 2+ secondary indications            → 4
    1 therapy area, 1 secondary (broad)                 → 3
    1 therapy area, 1 secondary (very niche)            → 2
    1 therapy area, primary indication(s) only          → 1
    """
    drug_rows = [r for r in enriched_rows
                 if (r.get("molecule_name") or "").lower() == molecule_name.lower()]
    if not drug_rows:
        return ""

    therapy_areas = set(
        r["therapy_area"] for r in drug_rows
        if r.get("therapy_area") and (r["therapy_area"] or "").lower() not in ("other", "")
    )
    secondary_indications = list(set(
        r["indication"] for r in drug_rows
        if (r.get("indication_type") or "").lower() == "secondary"
    ))
    primary_indications = list(set(
        r["indication"] for r in drug_rows
        if (r.get("indication_type") or "").lower() == "primary"
    ))

    if len(therapy_areas) > 1:
        return "5"
    elif len(secondary_indications) >= 2:
        return "4"
    elif len(secondary_indications) == 1:
        is_niche = _check_niche(molecule_name, primary_indications,
                                secondary_indications[0])
        return "2" if is_niche else "3"
    else:
        return "1"


def _check_niche(molecule, primary_list, secondary_indication):
    """Use Gemini + Search to determine if a secondary indication is very niche."""
    primary_str = ", ".join(primary_list) if primary_list else "unknown"
    prompt = f"""
You are a pharmaceutical analyst.

Drug: {molecule}
Primary indication(s): {primary_str}
Secondary (label expansion) indication: {secondary_indication}

Is this label expansion "very niche"?
A label expansion is "very niche" if it targets a narrow, specialized patient
sub-population, a rare disease, an orphan indication, or a highly specific
clinical context that affects relatively few patients compared to the primary
indication.

Search the web to determine the patient population size for this secondary
indication compared to the primary indication.

Return ONLY valid JSON:
{{"is_niche": true}} or {{"is_niche": false}}
No explanation.
"""
    try:
        text = _gemini_generate(
            gemini_client, types, prompt,
            system_instruction="Return ONLY valid JSON.",
            use_search=True,
        )
        if not text:
            return False
        data = _extract_json(text)
        return data.get("is_niche", False)
    except Exception as e:
        print(f"   ⚠️  Niche check failed: {e} — defaulting to non-niche")
        return False


# ══════════════════════════════════════════════════════════════════════════════
#  GEMINI ENRICHMENT — EXTRACTION ONLY (no classification here)
# ══════════════════════════════════════════════════════════════════════════════

def _enrich_single_trial(row, molecule_name, checkpoint, checkpoint_file, cp_lock):
    """Extract indications from a single trial. Classification happens later."""
    trial_id = row.get("trial_id")

    with cp_lock:
        if trial_id in checkpoint:
            cp = checkpoint[trial_id]
            return (trial_id, cp["conditions"], cp["trial_title"],
                    cp.get("phase", ""), row)

    prompt = f"""
You are a clinical trial data assistant.

Trial details:
Molecule: {row.get('molecule_name')}
Company: {row.get('company_name')}
Trial ID: {trial_id}
Phase: {row.get('phase')}
Source URL: {row.get('source_url')}

═══ STEP 1 — Look up the trial ═══
Search for this trial using the Trial ID "{trial_id}" on ClinicalTrials.gov
or other clinical trial registries (e.g. EudraCT, WHO ICTRP).
Find the EXACT official trial title as registered.

If the source URL is provided, also check that URL for the trial title.

═══ STEP 2 — Extract indications ═══
From the trial record, extract ALL disease indications being studied.
Include BOTH the primary indication AND any secondary/exploratory indications
that have documented outcomes.

Look at:
- The official trial title (most reliable source of the indication)
- The trial's "Conditions" or "Diseases" field on the registry
- Primary and secondary outcome measures
- The trial description / brief summary

Common patterns:
- "in subjects with type 2 diabetes" → T2DM
- "cardiovascular outcomes" → CV Risk Reduction
- Trial acronyms like PIONEER, SUSTAIN, STEP, SELECT often indicate specific programs
- Outcome studies (e.g., cardiovascular, renal) count as indications

Return ONLY valid JSON:
{{
  "conditions": [
    {{"indication": "<disease or condition>", "rationale": "<why — cite the trial record field>"}},
    {{"indication": "<disease or condition>", "rationale": "<why>"}}
  ],
  "trial_title": "<EXACT official trial title as registered on the clinical trial registry>",
  "phase": "<Phase from the registry, e.g. Phase 1, Phase 2, Phase 3, Phase 4, Phase 2/3>"
}}

Rules:
- trial_title must be the EXACT title from the registry, not a summary or guess
- phase must match what the registry lists (e.g. Phase 3, Phase 2/Phase 3)
- Include ALL indications the trial is evaluating
- Always extract at least the primary indication
- No explanations outside the JSON
"""
    try:
        import threading

        result_holder = {}
        error_holder  = {}

        def _call_gemini():
            try:
                text = _gemini_generate(
                    gemini_client, types, prompt,
                    system_instruction=(
                        "You are a clinical trial data assistant. "
                        "Search for the trial on ClinicalTrials.gov or other registries "
                        "to get the exact title. Return ONLY valid JSON."
                    ),
                    use_search=True,
                )
                result_holder["text"] = text
            except Exception as ex:
                error_holder["error"] = ex

        thread = threading.Thread(target=_call_gemini, daemon=True)
        thread.start()
        thread.join(timeout=90)

        if thread.is_alive():
            print(f"   ⏱️  Timeout (>90s) — skipping trial {trial_id}")
            conditions  = []
            trial_title = "Skipped (timeout)"
        elif "error" in error_holder:
            raise error_holder["error"]
        else:
            text = result_holder.get("text", "")
            data        = _extract_json(text)
            conditions  = data.get("conditions", [])
            trial_title = data.get("trial_title", "N/A")
            extracted_phase = (data.get("phase") or "").strip()

            if isinstance(conditions, list):
                normalized = []
                for c in conditions:
                    if isinstance(c, dict):
                        normalized.append({
                            "indication": (c.get("indication") or "").strip(),
                            "rationale":  (c.get("rationale") or "").strip(),
                        })
                    elif isinstance(c, str) and c.strip():
                        normalized.append({"indication": c.strip(), "rationale": ""})
                conditions = normalized
            else:
                conditions = [{"indication": str(conditions), "rationale": ""}]

            # Deduplicate
            seen, deduped = set(), []
            for c in conditions:
                key = (c["indication"] or "").lower()
                if key and key not in seen:
                    seen.add(key)
                    deduped.append(c)
            conditions = deduped

            print(f"   ✅ {trial_id}: {', '.join(c['indication'] for c in conditions) if conditions else 'no indications'}")

            with cp_lock:
                checkpoint[trial_id] = {"conditions": conditions, "trial_title": trial_title,
                                        "phase": extracted_phase}
                save_checkpoint(checkpoint_file, checkpoint)

    except Exception as e:
        conditions  = []
        trial_title = str(e)
        extracted_phase = ""
        print(f"   ❌ {trial_id}: {e}")

    return (trial_id, conditions, trial_title, extracted_phase, row)


def enrich_with_gemini(rows, molecule_name):
    import threading

    total = len(rows)
    checkpoint_file = f"checkpoint_{molecule_name.lower().replace(' ', '_')}.json"
    checkpoint = load_checkpoint(checkpoint_file)
    cp_lock = threading.Lock()
    print(f"📁 Loaded checkpoint → {len(checkpoint)} completed rows\n")
    print(f"🚀 Processing {total} trials with {MAX_WORKERS} parallel workers…\n")

    # ── STEP 1: Parallel extraction ──────────────────────────────────────
    results = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(
                _enrich_single_trial, row, molecule_name,
                checkpoint, checkpoint_file, cp_lock
            ): i
            for i, row in enumerate(rows)
        }
        for future in as_completed(futures):
            idx = futures[future]
            try:
                results.append(future.result())
            except Exception as e:
                row = rows[idx]
                print(f"   ❌ Unexpected error for {row.get('trial_id')}: {e}")
                results.append((row.get("trial_id"), [], str(e), "", row))

    # Sort by original order
    trial_order = {row.get("trial_id"): i for i, row in enumerate(rows)}
    results.sort(key=lambda r: trial_order.get(r[0], 0))

    # ── Unnest into flat rows (no classification yet) ────────────────────
    flat_rows = []
    for trial_id, conditions, trial_title, extracted_phase, row in results:
        # Use LLM-extracted phase as fallback when BigQuery phase is empty
        phase = row.get("phase") or extracted_phase or ""

        if not conditions:
            flat_rows.append({
                "molecule_name": row.get("molecule_name"),
                "company_name":  row.get("company_name"),
                "indication":    "No indication found",
                "rationale":     "",
                "trial_title":   trial_title,
                "trial_id":      trial_id,
                "phase":         phase,
                "source_url":    row.get("source_url"),
                "data_source":   "Clinical Trials",
            })
            continue

        seen = set()
        for c in conditions:
            raw = c.get("indication", "")
            if not raw:
                continue
            std = standardize_indication(raw)
            if not std or std.lower() in ("error", "n/a", "no indication found", "none"):
                continue
            if std.lower() in seen:
                continue
            seen.add(std.lower())

            flat_rows.append({
                "molecule_name": row.get("molecule_name"),
                "company_name":  row.get("company_name"),
                "indication":    std,
                "rationale":     c.get("rationale", ""),
                "trial_title":   trial_title,
                "trial_id":      trial_id,
                "phase":         phase,
                "source_url":    row.get("source_url"),
                "data_source":   "Clinical Trials",
            })

    # ── STEP 2: Classify all unique indications in ONE LLM+Search call ──
    unique_indications = sorted(set(
        r["indication"] for r in flat_rows
        if r["indication"] != "No indication found"
    ))

    classification_map = classify_indications_with_llm(molecule_name, unique_indications)

    # Apply classification to every row
    for row in flat_rows:
        ind_key = (row["indication"] or "").lower()
        cls = classification_map.get(ind_key, {})
        row["indication_type"] = cls.get("indication_type", "")
        row["therapy_area"]    = cls.get("therapy_area", "")
        row["rationale"]       = cls.get("rationale", "")

    # Override: Primary + non-Metabolic therapy area → Secondary
    for row in flat_rows:
        if ((row.get("indication_type") or "").lower() == "primary"
                and (row.get("therapy_area") or "").lower() != "metabolic"):
            row["indication_type"] = "Secondary"

    # ── STEP 3: Compute Ep (row-wise, only for Primary/Secondary) ──────
    for row in flat_rows:
        ind_type = (row.get("indication_type") or "").lower()
        if ind_type in ("primary", "secondary"):
            row["Ep"] = compute_ep(row.get("phase"))
        else:
            row["Ep"] = ""

    # ── STEP 4: Compute Et (drug-level, only Primary/Secondary rows) ──
    print("🔹 Computing Et (drug-level expansion score)…")
    scored_rows = [r for r in flat_rows
                   if (r.get("indication_type") or "").lower() in ("primary", "secondary")]
    et_value = compute_et(scored_rows, molecule_name)
    for row in flat_rows:
        row["Et"] = et_value
    print(f"   Et = {et_value}")

    print(f"\n✅ Completed. {len(flat_rows)} rows (extracted → classified → scored)\n")
    return flat_rows


# ── COMMON FORMAT HEADERS ────────────────────────────────────────────────────
HEADERS    = ["molecule_name", "company_name", "indication", "rationale",
              "indication_type", "therapy_area", "trial_title", "trial_id",
              "phase", "source_url", "data_source", "Ep", "Et"]
COL_WIDTHS = [18, 22, 28, 40, 16, 18, 50, 18, 10, 40, 16, 8, 8]


def _thin_border():
    return Border(bottom=Side(style="thin", color="DDDDDD"),
                  right=Side(style="thin", color="DDDDDD"))


def _write_summary_sheet(wb, rows, title_label="Summary"):
    """Add a 'Summary' sheet: Drug Name | No. of Indications | No. of Therapy Areas."""
    ws = wb.create_sheet("Summary")
    thin  = _thin_border()
    hfill = PatternFill("solid", start_color="1A1A2E")
    alt   = PatternFill("solid", start_color="F4F7FB")

    # Aggregate per drug
    from collections import defaultdict
    drug_data = defaultdict(lambda: {"indications": set(), "therapy_areas": set()})
    for r in rows:
        name = r.get("molecule_name", "")
        ind  = r.get("indication", "")
        ta   = r.get("therapy_area", "")
        if not name or ind == "No indication found":
            continue
        drug_data[name]["indications"].add(ind)
        if ta and ta.lower() not in ("other", ""):
            drug_data[name]["therapy_areas"].add(ta)

    summary_headers = ["Drug Name", "No. of Indications", "No. of Therapy Areas"]
    summary_widths  = [28, 20, 22]

    # Title row
    ws.merge_cells("A1:C1")
    c           = ws["A1"]
    c.value     = f"Summary  —  {title_label}"
    c.font      = Font(name="Arial", bold=True, size=14, color="1A1A2E")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    # Header row
    for col, h in enumerate(summary_headers, 1):
        c           = ws.cell(row=2, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = hfill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin
    ws.row_dimensions[2].height = 22

    # Data rows
    for idx, (drug, info) in enumerate(sorted(drug_data.items()), start=3):
        fill = alt if idx % 2 == 1 else None
        vals = [drug, len(info["indications"]), len(info["therapy_areas"])]
        for col, val in enumerate(vals, 1):
            c           = ws.cell(row=idx, column=col, value=val)
            c.font      = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal="center" if col > 1 else "left",
                                    vertical="center")
            c.border    = thin
            if fill:
                c.fill = fill
        ws.row_dimensions[idx].height = 20

    for i, w in enumerate(summary_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_excel(rows, molecule_name, output_dir=None):
    print("🔹 Writing Excel...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clinical Efficacy"

    thin     = _thin_border()
    hfill    = PatternFill("solid", start_color="1A1A2E")
    alt_fill = PatternFill("solid", start_color="F4F7FB")
    ncols    = len(HEADERS)

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value     = f"Clinical Efficacy  —  {molecule_name.title()}"
    c.font      = Font(name="Arial", bold=True, size=14, color="1A1A2E")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    # Header row
    for col, h in enumerate(HEADERS, 1):
        c           = ws.cell(row=2, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = hfill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin
    ws.row_dimensions[2].height = 22

    # Data rows
    for idx, row in enumerate(rows, start=3):
        fill = alt_fill if idx % 2 == 1 else None
        for col, key in enumerate(HEADERS, 1):
            c           = ws.cell(row=idx, column=col, value=row.get(key, ""))
            c.font      = Font(name="Arial", size=9)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border    = thin
            if fill:
                c.fill = fill
        ws.row_dimensions[idx].height = 18

    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}{2 + len(rows)}"
    ws.freeze_panes = "A3"
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Summary sheet ────────────────────────────────────────────────────
    _write_summary_sheet(wb, rows, molecule_name)

    file_name = f"{molecule_name.lower().replace(' ', '_')}_clinical_efficacy.xlsx"
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        file_name = os.path.join(output_dir, file_name)
    wb.save(file_name)
    print(f"✅ Excel saved → {file_name}\n")
    return file_name


def main():
    from datetime import datetime

    parser = argparse.ArgumentParser()
    parser.add_argument("molecule_name")
    args = parser.parse_args([a for a in sys.argv[1:] if a != "--"])
    molecule = args.molecule_name

    # Create timestamped output folder
    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = f"output_{timestamp}"
    os.makedirs(output_dir, exist_ok=True)

    print("\n🚀 Pipeline Started\n")
    print(f"📂 Output folder: {output_dir}\n")
    print("[1/3] Fetching data...")
    rows = fetch_rows(molecule)
    if not rows:
        print("❌ No data found")
        return
    print("[2/3] Enriching data...")
    enriched = enrich_with_gemini(rows, molecule)
    print("[3/3] Generating Excel...")
    output = write_excel(enriched, molecule, output_dir)
    print("🎉 DONE")
    print(f"📄 File: {output}")
    print(f"📊 Rows processed: {len(enriched)}")

if __name__ == "__main__":
    main()
