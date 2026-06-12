"""
Drug Indication Researcher
===========================
Researches drug indications EXCLUSIVELY from innovator / investor sources:
  - Company investor presentations & R&D day slides
  - Company pipeline pages
  - FDA / EMA regulatory labels (by the innovator)
  - Press releases & earnings calls by the innovator company
  - SEC filings (10-K, 10-Q, 8-K, 20-F, 6-K) by the innovator company

Does NOT use BigQuery or clinical trial databases.
Uses Gemini 2.5 Flash with Google Search grounding.

Key difference from label.py (clinical trials): this file captures what the
INNOVATOR COMPANY says about the drug in their own materials.
Each row = one indication × one source document.

Output: one indication per row in standardized format.

Usage:
    python drug_indication_researcher.py semaglutide
    python drug_indication_researcher.py semaglutide --company "Novo Nordisk"
"""

import sys
import os
import json
import re
import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from dotenv import load_dotenv
from google import genai
from google.genai import types
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from indication_standardizer import standardize_indication, classify_indication, get_therapy_area, process_indications

load_dotenv(override=True)

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
gemini_client  = genai.Client(api_key=GEMINI_API_KEY)

# ── Parallel config ───────────────────────────────────────────────────────────
MAX_WORKERS = 5  # parallel Gemini calls for innovator source queries

# ══════════════════════════════════════════════════════════════════════════════
#  SECONDARY INDICATION CRITERIA PROMPT
#  Applied when extracting indications from any source document.
# ══════════════════════════════════════════════════════════════════════════════

SECONDARY_INDICATION_CRITERIA = """
You must extract ONLY secondary indications. Do NOT include the primary indication.

A secondary indication qualifies ONLY if ALL of the following are true:

- The indication represents a true expansion — i.e., it is not part of the primary indication (for clinical assets) or currently approved label (for commercial assets)
- The indication is described at a clear disease-level definition, avoiding vague, overlapping, or synonymous representations
- The source must describe observed or measured outcomes in that specific indication (e.g., trial results, endpoint readouts, biomarker response), not just planned evaluation or exploratory intent

The following must NOT be considered secondary indications:
* Indications mentioned only as hypothesis, targets, or exploratory possibilities
* Pipeline indications without any data or outcomes
* Mechanism based assumptions without clinical or empirical validation
* Early discovery or preclinical signals without human data
* Any indication lacking traceable, verifiable evidence of results

If no indication meets these criteria, return an empty list [].
"""


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 1 — Identify the innovator company (if not provided)
# ══════════════════════════════════════════════════════════════════════════════

def identify_company(molecule: str) -> str:
    """Use Gemini + Search to find the innovator/originator company."""
    prompt = (
        f"Who is the innovator (originator) pharmaceutical company that developed {molecule}? "
        f"Return ONLY a JSON object: {{\"company\": \"<company name>\", \"brand_names\": [\"name1\", \"name2\"]}}\n"
        f"No explanation."
    )
    try:
        resp = gemini_client.models.generate_content(
            model="gemini-2.5-flash",
            contents=prompt,
            config=types.GenerateContentConfig(
                temperature=0,
                tools=[types.Tool(google_search=types.GoogleSearch())],
                system_instruction="Return ONLY valid JSON. No markdown fences.",
            ),
        )
        text = resp.text.strip()
        text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text).strip()
        data = json.loads(text)
        company = data.get("company", "")
        brands  = data.get("brand_names", [])
        print(f"  🏢 Innovator   : {company}")
        if brands:
            print(f"  💊 Brand names : {', '.join(brands)}")
        return company
    except Exception as e:
        print(f"  ⚠️  Could not identify company: {e}")
        return ""


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 2 — Research queries, each targeting a specific innovator source type
# ══════════════════════════════════════════════════════════════════════════════

def _build_queries(molecule: str, company: str) -> list[dict]:
    return [
        # ── 1. FDA / EMA labels ───────────────────────────────────────────
        {
            "source_type": "Regulatory Label",
            "prompt": (
                f"Search the web for official FDA prescribing information and EMA SmPC "
                f"for {molecule} (by {company}).\n\n"
                f"For EACH approved indication found in those labels, return a separate entry.\n\n"
                f"Return ONLY valid JSON:\n"
                f'{{"entries": [\n'
                f'  {{\n'
                f'    "indication": "<disease or condition>",\n'
                f'    "brand_name": "<commercial brand name, e.g. Ozempic, Wegovy>",\n'
                f'    "source_document": "<exact label title, e.g. Ozempic FDA Prescribing Information>",\n'
                f'    "source_url": "<URL to the label or DailyMed page>",\n'
                f'    "detail": "<approval year, patient population, dose>",\n'
                f'    "rationale": "<why this indication was identified — cite the specific label section or approval>"\n'
                f"  }}\n"
                f"]}}\n\n"
                f"IMPORTANT:\n"
                f"- One entry per indication, not per brand\n"
                f"- If the drug has multiple brands for different indications, list each separately\n"
                f"- source_document must be the ACTUAL document title, not a generic description\n"
                f"- No explanation, only JSON"
            ),
        },
        # ── 2. Investor presentations / R&D day ──────────────────────────
        {
            "source_type": "Investor Presentation",
            "prompt": (
                f"Search the web for {company}'s latest investor presentations, "
                f"Capital Markets Day slides, R&D day presentations, and annual report "
                f"that mention {molecule}.\n\n"
                f"For EACH indication mentioned for {molecule} in these presentations, "
                f"return a separate entry.\n\n"
                f"{SECONDARY_INDICATION_CRITERIA}\n"
                f"Return ONLY valid JSON:\n"
                f'{{"entries": [\n'
                f'  {{\n'
                f'    "indication": "<disease or condition>",\n'
                f'    "brand_name": "<brand name if mentioned>",\n'
                f'    "source_document": "<exact presentation title, e.g. Novo Nordisk Capital Markets Day 2024>",\n'
                f'    "source_url": "<URL to the presentation PDF or page>",\n'
                f'    "detail": "<specific claim, trial name, or pipeline status from the slide>",\n'
                f'    "rationale": "<why this indication was identified — cite the specific evidence from the presentation>"\n'
                f"  }}\n"
                f"]}}\n\n"
                f"IMPORTANT:\n"
                f"- source_document must be the REAL title of the presentation, not a generic label\n"
                f"- Include the year in the source_document title\n"
                f"- detail should quote or paraphrase specific claims from the presentation\n"
                f"- Only include indications with observed outcomes data — not exploratory mentions\n"
                f"- No explanation, only JSON"
            ),
        },
        # ── 3. Press releases / earnings calls ───────────────────────────
        {
            "source_type": "Press Release",
            "prompt": (
                f"Search the web for press releases and earnings call statements "
                f"from {company} about {molecule}.\n\n"
                f"For EACH indication mentioned in these press releases or earnings calls, "
                f"return a separate entry.\n\n"
                f"{SECONDARY_INDICATION_CRITERIA}\n"
                f"Return ONLY valid JSON:\n"
                f'{{"entries": [\n'
                f'  {{\n'
                f'    "indication": "<disease or condition>",\n'
                f'    "brand_name": "<brand name if mentioned>",\n'
                f'    "source_document": "<exact press release title or earnings call date>",\n'
                f'    "source_url": "<URL to the press release>",\n'
                f'    "detail": "<key announcement: approval, trial result, filing, etc.>",\n'
                f'    "rationale": "<why this indication was identified — cite the specific news or data>"\n'
                f"  }}\n"
                f"]}}\n\n"
                f"IMPORTANT:\n"
                f"- source_document must be the REAL press release headline or earnings call date\n"
                f"- detail should include the specific news (e.g. FDA approval, Phase 3 results)\n"
                f"- Only include indications backed by reported outcomes, not future plans\n"
                f"- No explanation, only JSON"
            ),
        },
        # ── 4. Company pipeline page ─────────────────────────────────────
        {
            "source_type": "Pipeline Page",
            "prompt": (
                f"Search the web for {company}'s official pipeline page or product portfolio page "
                f"that lists {molecule}.\n\n"
                f"For EACH indication listed for {molecule} on the pipeline, return a separate entry.\n\n"
                f"{SECONDARY_INDICATION_CRITERIA}\n"
                f"Return ONLY valid JSON:\n"
                f'{{"entries": [\n'
                f'  {{\n'
                f'    "indication": "<disease or condition>",\n'
                f'    "brand_name": "<brand name if shown>",\n'
                f'    "source_document": "<e.g. Novo Nordisk Pipeline — as of Q1 2025>",\n'
                f'    "source_url": "<URL to the pipeline page>",\n'
                f'    "detail": "<any additional status note from the pipeline>",\n'
                f'    "rationale": "<why this indication was identified — cite the pipeline entry>"\n'
                f"  }}\n"
                f"]}}\n\n"
                f"IMPORTANT:\n"
                f"- One entry per indication\n"
                f"- Exclude Preclinical and early discovery entries (no human data)\n"
                f"- No explanation, only JSON"
            ),
        },
        # ── 5. SEC filings (10-K, 10-Q, 8-K, 20-F, 6-K) ─────────────────
        {
            "source_type": "SEC Filing",
            "prompt": (
                f"Search the SEC EDGAR database and the web for SEC filings by {company} "
                f"(10-K annual reports, 10-Q quarterly reports, 8-K current reports, "
                f"20-F annual reports for foreign private issuers, 6-K reports) "
                f"that disclose clinical or regulatory information about {molecule}.\n\n"
                f"For EACH indication described with actual clinical data or regulatory outcomes "
                f"in these filings, return a separate entry.\n\n"
                f"{SECONDARY_INDICATION_CRITERIA}\n"
                f"Return ONLY valid JSON:\n"
                f'{{"entries": [\n'
                f'  {{\n'
                f'    "indication": "<disease or condition>",\n'
                f'    "brand_name": "<brand name if mentioned>",\n'
                f'    "source_document": "<exact filing type and period, e.g. {company} 10-K FY2024, {company} 8-K dated 2024-03-15>",\n'
                f'    "source_url": "<URL to the SEC EDGAR filing or press release>",\n'
                f'    "detail": "<specific disclosed data: trial outcome, milestone, risk factor, MD&A discussion>",\n'
                f'    "rationale": "<why this indication was identified — cite the filing section and data>"\n'
                f"  }}\n"
                f"]}}\n\n"
                f"IMPORTANT:\n"
                f"- source_document must identify the exact filing type, company, and reporting period\n"
                f"- Only include indications where the filing reports actual results or regulatory "
                f"  decisions — not forward-looking statements or risk factor boilerplate\n"
                f"- detail must reference the specific section of the filing (e.g. 'Business — Products', "
                f"  'MD&A', 'Risk Factors', 'Item 8-K Exhibit 99.1') where the data appears\n"
                f"- Prioritize 8-K filings that announce trial readouts, FDA approvals, or NDA/BLA submissions\n"
                f"- No explanation, only JSON"
            ),
        },
    ]


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 3 — Run research + build rows (PARALLELIZED)
# ══════════════════════════════════════════════════════════════════════════════

def _research_single_source(q, molecule, company):
    """Query a single innovator source type. Returns list of row dicts."""
    source_type = q["source_type"]
    prompt      = q["prompt"]
    rows = []
    print(f"\n  🔍 {source_type}…")

    try:
        response = gemini_client.models.generate_content(
            model="gemini-2.5-flash",
            contents=prompt,
            config=types.GenerateContentConfig(
                temperature=0,
                tools=[types.Tool(google_search=types.GoogleSearch())],
                system_instruction=(
                    "You are a pharmaceutical analyst researching what the "
                    "innovator company publicly says about this drug. "
                    "Search the web and SEC EDGAR. Return ONLY valid JSON — no markdown, "
                    "no explanation, no commentary. "
                    "Apply strict secondary indication criteria: only include indications "
                    "with documented, verifiable clinical outcomes."
                ),
            ),
        )

        text = response.text.strip()
        text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text).strip()
        data    = json.loads(text)
        entries = data.get("entries", [])

        if not entries:
            print(f"     ⚠️  No entries found")
            return rows

        print(f"     ✅ {len(entries)} entries")

        for e in entries:
            raw_indication = e.get("indication", "").strip()
            if not raw_indication:
                continue

            std_indication = standardize_indication(raw_indication)
            if std_indication.lower() in ("error", "n/a", "none", "no indication found"):
                continue

            ind_type = classify_indication(molecule, std_indication)

            # Build rationale: prefer explicit rationale, fall back to detail
            rationale = (e.get("rationale") or e.get("detail") or "").strip()

            rows.append({
                "molecule_name":   molecule.title(),
                "company_name":    company,
                "indication":      std_indication,
                "rationale":       rationale,
                "indication_type": ind_type,
                "therapy_area":    get_therapy_area(std_indication),
                "trial_title":     e.get("source_document", ""),
                "trial_id":        e.get("brand_name", ""),
                "phase":           "",
                "source_url":      e.get("source_url", ""),
                "data_source":     f"Innovator: {source_type}",
            })
            print(f"       • {std_indication} ({ind_type}) ← {e.get('source_document', '?')[:50]}")

    except json.JSONDecodeError:
        print(f"     ⚠️  JSON parse failed — trying fallback")
        _fallback_extract(molecule, company, source_type, text, rows)
    except Exception as e:
        print(f"     ❌ Error: {e}")

    return rows


def research_indications(molecule: str, company: str) -> list[dict]:
    queries  = _build_queries(molecule, company)
    all_rows = []

    print(f"\n  🚀 Querying {len(queries)} innovator sources in parallel…\n")

    # ── Parallel source queries ──────────────────────────────────────────
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(_research_single_source, q, molecule, company): q["source_type"]
            for q in queries
        }
        for future in as_completed(futures):
            source_type = futures[future]
            try:
                rows = future.result()
                all_rows.extend(rows)
            except Exception as ex:
                print(f"     ❌ {source_type}: {ex}")

    # ── Deduplicate only exact duplicates (same indication + same source doc)
    seen   = set()
    unique = []
    for row in all_rows:
        key = (row["indication"], row["trial_title"], row["data_source"])
        if key in seen:
            continue
        seen.add(key)
        unique.append(row)

    print(f"\n  📊 Total rows: {len(unique)} (deduped from {len(all_rows)})")
    return unique


def _fallback_extract(molecule, company, source_type, raw_text, all_rows):
    """Second-chance extraction if JSON parse failed."""
    try:
        prompt = (
            f"The following text is about {molecule} by {company}.\n"
            f"Extract every disease indication mentioned that has documented clinical outcomes.\n"
            f"Return ONLY a JSON array: [\"indication1\", \"indication2\"]\n\n"
            f"Strict rules — DO NOT include:\n"
            f"- Hypothetical or exploratory mentions without outcome data\n"
            f"- Preclinical or mechanism-based claims\n"
            f"- Pipeline entries with no trial results reported\n\n"
            f"Text:\n{raw_text[:3000]}"
        )
        resp = gemini_client.models.generate_content(
            model="gemini-2.5-flash",
            contents=prompt,
            config=types.GenerateContentConfig(
                temperature=0,
                system_instruction="Return ONLY a JSON array of strings.",
            ),
        )
        text = resp.text.strip()
        text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text).strip()
        parsed = json.loads(text)
        if isinstance(parsed, list):
            for raw in parsed:
                std = standardize_indication(str(raw))
                if std.lower() in ("error", "n/a", "none"):
                    continue
                all_rows.append({
                    "molecule_name":   molecule.title(),
                    "company_name":    company,
                    "indication":      std,
                    "rationale":       "(extracted from unstructured response via fallback)",
                    "indication_type": classify_indication(molecule, std),
                    "therapy_area":    get_therapy_area(std),
                    "trial_title":     "(extracted from unstructured response)",
                    "trial_id":        "",
                    "phase":           "",
                    "source_url":      "",
                    "data_source":     f"Innovator: {source_type}",
                })
            print(f"     🔄 Fallback: {len(parsed)} indications")
    except Exception:
        print(f"     ⚠️  Fallback also failed")


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL — same column format as label.py
# ══════════════════════════════════════════════════════════════════════════════

HEADERS    = ["molecule_name", "company_name", "indication", "rationale",
              "indication_type", "therapy_area", "trial_title", "trial_id",
              "phase", "source_url", "data_source"]
COL_WIDTHS = [18, 22, 28, 40, 16, 18, 50, 18, 10, 40, 16]


def _thin_border():
    return Border(bottom=Side(style="thin", color="DDDDDD"),
                  right=Side(style="thin", color="DDDDDD"))


def write_excel(rows, molecule_name):
    print("🔹 Writing Excel…")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Innovator Research"

    thin     = _thin_border()
    hfill    = PatternFill("solid", start_color="1A1A2E")
    alt_fill = PatternFill("solid", start_color="F4F7FB")
    ncols    = len(HEADERS)

    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value     = f"Drug Indication Research (Innovator Sources)  —  {molecule_name.title()}"
    c.font      = Font(name="Arial", bold=True, size=14, color="1A1A2E")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    for col, h in enumerate(HEADERS, 1):
        c           = ws.cell(row=2, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = hfill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin
    ws.row_dimensions[2].height = 22

    for idx, row in enumerate(rows, start=3):
        fill = alt_fill if idx % 2 == 1 else None
        for col, key in enumerate(HEADERS, 1):
            c           = ws.cell(row=idx, column=col, value=row.get(key, ""))
            c.font      = Font(name="Arial", size=9)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border    = thin
            if fill:
                c.fill = fill
        ws.row_dimensions[idx].height = 18

    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}{2 + len(rows)}"
    ws.freeze_panes = "A3"
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    file_name = f"{molecule_name.lower().replace(' ', '_')}_indication_research.xlsx"
    wb.save(file_name)
    print(f"✅ Excel saved → {file_name}\n")
    return file_name


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Research drug indications from innovator/investor/SEC sources (Gemini + Google Search)"
    )
    parser.add_argument("molecule_name", help="Drug/molecule name, e.g. semaglutide")
    parser.add_argument("--company", default=None,
                        help="Innovator company name (auto-detected if not provided)")
    args = parser.parse_args([a for a in sys.argv[1:] if a != "--"])

    molecule = args.molecule_name
    company  = args.company

    if not GEMINI_API_KEY:
        sys.exit("❌  GEMINI_API_KEY not set in .env")

    print(f"\n🚀 Drug Indication Researcher (Innovator + SEC Sources)")
    print(f"   Molecule : {molecule.title()}")

    # Auto-detect company if not provided
    if not company:
        print("\n  🔎 Identifying innovator company…")
        company = identify_company(molecule)
        if not company:
            print("  ⚠️  Could not detect company. Provide --company for better results.")
            company = ""
    else:
        print(f"   Company  : {company}")

    print(f"\n[1/2] Researching innovator + SEC sources for {molecule.title()} ({company})…")
    rows = research_indications(molecule, company)

    if not rows:
        print("❌ No indications found from innovator/SEC sources")
        return

    print(f"\n[2/2] Generating Excel…")
    output = write_excel(rows, molecule)

    print("🎉 DONE")
    print(f"📄 File: {output}")
    print(f"📊 Rows: {len(rows)}")


if __name__ == "__main__":
    main()
