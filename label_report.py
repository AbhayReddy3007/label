#!/usr/bin/env python3
"""
Label Expansion Report Generator (Business-Facing)
====================================================
Reads the Excel files produced by run_all.py / label.py and generates
a concise, insight-driven PDF report (max 2 pages) for senior business
stakeholders.

The report focuses on indication breadth, label expansion potential,
therapy area coverage, and strategic implications — NOT scores.

Gemini 2.5 Flash generates the analytical narrative. Falls back to a
structured summary if the API is unavailable.

Usage:
    python label_report.py --molecule semaglutide
    python label_report.py --molecule semaglutide tirzepatide
    python label_report.py --output-dir output --molecule dupilumab
"""

import argparse
import glob
import json
import os
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_JUSTIFY
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, KeepTogether
)

import openpyxl

load_dotenv()

DEFAULT_OUTPUT_DIR = "output"

GEMINI_MODEL = "gemini-2.5-flash"
GEMINI_API_URL = (
    "https://generativelanguage.googleapis.com/v1beta/models/"
    f"{GEMINI_MODEL}:generateContent"
)


# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------

def normalize(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


def find_excel_file(output_dir: str, molecule: str) -> str | None:
    """Locate the Excel file produced by run_all.py or label.py for a molecule."""
    safe = re.sub(r"[^a-zA-Z0-9]", "_", molecule.strip().lower())

    # Patterns: <molecule>_research.xlsx (run_all.py) or <molecule>_clinical_efficacy.xlsx (label.py)
    patterns = [
        str(Path(output_dir) / f"{safe}_research.xlsx"),
        str(Path(output_dir) / f"{safe}_clinical_efficacy.xlsx"),
        str(Path(output_dir) / f"*{safe}*.xlsx"),
    ]
    for pattern in patterns:
        matches = glob.glob(pattern)
        if matches:
            matches.sort(key=lambda p: os.path.getmtime(p), reverse=True)
            return matches[0]

    # Broad search
    all_xlsx = glob.glob(str(Path(output_dir) / "*.xlsx"))
    for f in all_xlsx:
        if normalize(molecule) in normalize(Path(f).stem):
            return f
    return None


def read_excel_data(file_path: str) -> list[dict]:
    """Read rows from the 'All Combined' or 'Clinical Efficacy' sheet."""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    # Prefer 'All Combined' sheet, fall back to first sheet
    for sheet_name in ["All Combined", "Clinical Efficacy", "All Molecules"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            break
    else:
        ws = wb.active

    rows = []
    headers = None
    for row_cells in ws.iter_rows(values_only=True):
        vals = list(row_cells)
        # Skip title rows (merged cells with a single value)
        non_empty = [v for v in vals if v is not None and str(v).strip()]
        if len(non_empty) <= 2:
            continue

        # Detect header row
        if headers is None:
            lower_vals = [str(v).strip().lower() for v in vals if v]
            if "indication" in lower_vals or "molecule_name" in lower_vals:
                headers = [str(v).strip() if v else f"col_{i}" for i, v in enumerate(vals)]
                continue
            continue

        row_dict = {}
        for i, h in enumerate(headers):
            row_dict[h] = vals[i] if i < len(vals) else ""
        rows.append(row_dict)

    wb.close()
    return rows


def extract_label_data(rows: list[dict], molecule: str) -> dict:
    """Extract structured data from the Excel rows for report generation."""
    # Filter rows for this molecule
    mol_rows = [
        r for r in rows
        if normalize(r.get("molecule_name", "")) == normalize(molecule)
        or normalize(molecule) in normalize(r.get("molecule_name", ""))
    ]
    if not mol_rows:
        mol_rows = rows  # Use all rows if molecule filter returns nothing

    # Gather indications (exclude "No indication found")
    all_indications = set()
    primary_indications = set()
    secondary_indications = set()
    secondary_indication_rationales: dict[str, dict] = {}  # ind → {rationale, therapy_area, phase}
    therapy_areas = set()
    data_sources = set()
    trials = set()
    phases = defaultdict(int)
    indication_details = []  # For passing to Gemini

    company_name = ""
    has_regulatory_label = False

    for r in mol_rows:
        ind = (r.get("indication") or "").strip()
        if not ind or ind.lower() in ("no indication found", "n/a", "none", "error"):
            continue

        ind_type = (r.get("indication_type") or "").strip().lower()
        if ind_type not in ("primary", "secondary"):
            continue

        all_indications.add(ind)

        if ind_type == "primary":
            primary_indications.add(ind)
        elif ind_type == "secondary":
            secondary_indications.add(ind)
            # Keep the longest / most informative rationale per indication
            rationale = (r.get("rationale") or "").strip()
            existing = secondary_indication_rationales.get(ind)
            if not existing or len(rationale) > len(existing.get("rationale", "")):
                secondary_indication_rationales[ind] = {
                    "rationale": rationale,
                    "therapy_area": (r.get("therapy_area") or "").strip(),
                    "phase": (r.get("phase") or "").strip(),
                }

        ta = (r.get("therapy_area") or "").strip()
        if ta and ta.lower() not in ("other", ""):
            therapy_areas.add(ta)

        ds = (r.get("data_source") or "").strip()
        if ds:
            data_sources.add(ds)
            if "regulatory" in ds.lower() or "label" in ds.lower():
                has_regulatory_label = True

        tid = (r.get("trial_id") or "").strip()
        if tid:
            trials.add(tid)

        phase = (r.get("phase") or "").strip()
        if phase:
            phases[phase] += 1

        if not company_name:
            company_name = (r.get("company_name") or "").strip()

        indication_details.append({
            "indication": ind,
            "type": r.get("indication_type", ""),
            "therapy_area": ta,
            "rationale": (r.get("rationale") or "")[:200],
            "trial_title": (r.get("trial_title") or "")[:100],
            "trial_id": tid,
            "phase": phase,
            "data_source": ds,
        })

    # Deduplicate indication_details by (indication, data_source)
    seen = set()
    unique_details = []
    for d in indication_details:
        key = (d["indication"].lower(), d["data_source"])
        if key not in seen:
            seen.add(key)
            unique_details.append(d)

    return {
        "molecule_name": mol_rows[0].get("molecule_name", molecule) if mol_rows else molecule,
        "company_name": company_name,
        "total_indications": len(all_indications),
        "primary_indications": sorted(primary_indications),
        "secondary_indications": sorted(secondary_indications),
        "secondary_indication_rationales": secondary_indication_rationales,
        "therapy_areas": sorted(therapy_areas),
        "data_sources": sorted(data_sources),
        "total_trials": len(trials),
        "phase_distribution": dict(phases),
        "has_regulatory_label": has_regulatory_label,
        "total_rows": len(mol_rows),
        "indication_details": unique_details[:40],  # Cap for prompt length
    }


def safe_text(val) -> str:
    if val is None:
        return "N/A"
    s = str(val).strip()
    return s if s else "N/A"


def escape_html(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def lookup_innovator(molecule_name: str, sponsor_hint: str = "") -> str | None:
    import urllib.request
    import urllib.error

    api_key = os.environ.get("GEMINI_API_KEY", "").strip()
    if not api_key:
        return None

    prompt = (
        f"Identify the original innovator company that first developed or "
        f"discovered the pharmaceutical molecule \"{molecule_name}\". "
        f"The company on record is \"{sponsor_hint or 'Unknown'}\". "
        f"Return ONLY the company name on a single line — no explanation, "
        f"no punctuation beyond what is in the name itself. "
        f"If you are unsure, return the company name as-is."
    )

    payload = json.dumps({
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.0, "maxOutputTokens": 64},
    }).encode("utf-8")

    url = f"{GEMINI_API_URL}?key={api_key}"
    req = urllib.request.Request(
        url, data=payload,
        headers={"Content-Type": "application/json"},
        method="POST",
    )

    try:
        print(f"  [AI]  Looking up innovator for {molecule_name}...")
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = json.loads(resp.read().decode("utf-8"))
        text = (
            result.get("candidates", [{}])[0]
                  .get("content", {})
                  .get("parts", [{}])[0]
                  .get("text", "")
        ).strip()
        if text:
            print(f"  [AI]  Innovator resolved: {text}")
            return text
        return None
    except Exception as e:
        print(f"  [WARN] Innovator lookup failed ({e}), using company as fallback.",
              file=sys.stderr)
        return None


# ---------------------------------------------------------------------
# Gemini narrative — business-facing analytical report
# ---------------------------------------------------------------------

def build_gemini_prompt(molecule_name: str, data: dict) -> str:
    """
    Build a prompt that instructs Gemini to produce the business-facing
    analytical report for the Label Expansion dimension.
    """
    # Build indication summary
    indication_lines = []
    for d in data["indication_details"][:30]:
        indication_lines.append(
            f"  {d['indication']} | Type: {d['type']} | "
            f"Therapy Area: {d['therapy_area']} | Phase: {d['phase']} | "
            f"Source: {d['data_source']}"
        )

    primary_list = ", ".join(data["primary_indications"][:10]) or "None identified"
    secondary_list = ", ".join(data["secondary_indications"][:15]) or "None identified"
    ta_list = ", ".join(data["therapy_areas"]) or "None identified"
    sources_list = ", ".join(data["data_sources"]) or "N/A"

    # Build secondary indication rationale lines
    rationales_map = data.get("secondary_indication_rationales", {})
    secondary_rationale_lines = []
    for ind in data["secondary_indications"][:15]:
        info = rationales_map.get(ind, {})
        rationale = info.get("rationale", "No rationale available")
        ta = info.get("therapy_area", "")
        phase = info.get("phase", "")
        line = f"  {ind}"
        if ta:
            line += f" | Therapy Area: {ta}"
        if phase:
            line += f" | {phase}"
        line += f"\n    Reason it is a secondary indication: {rationale}"
        secondary_rationale_lines.append(line)

    prompt = f"""You are a senior business analyst preparing a detailed analytical report 
on the label expansion dimension of the pharmaceutical molecule "{molecule_name}" 
for senior business decision-makers.

This dimension evaluates the breadth and depth of the drug's indication landscape — 
how many distinct indications it is approved for or being studied in, whether it has 
expanded beyond its primary use, how many therapy areas it spans, and what this means 
for the drug's commercial and strategic potential.

YOUR OUTPUT MUST FOLLOW THIS EXACT STRUCTURE (use these exact headings):

## HEADLINE
Write ONE impactful sentence summarizing the overall business implication of the 
label expansion landscape for this molecule. This should be the single most important 
takeaway a decision-maker needs.

## INDICATION LANDSCAPE
Write 3-5 sentences providing the quantitative context a decision-maker needs. 
Cover: how many distinct indications exist (primary vs. secondary/expansion), 
how many therapy areas the drug spans, what the primary indication(s) are and 
what expansion indications are being pursued, what data sources corroborate the 
findings (clinical trials, regulatory labels, investor materials, SEC filings), 
and the phase maturity of indication-level evidence. Cite specific numbers from 
the data. This section sets the stage — it should tell the reader the size and 
shape of the indication portfolio before diving into insights.

## KEY INSIGHTS

Provide 4-6 key insights. For EACH insight, write EXACTLY two lines using this format:
Line 1: "Insight: " followed by a short, specific finding (ONE sentence, max 20 words).
         This is the bold headline of the insight.
Line 2: "Why it matters: " followed by the business implication (2-3 sentences, 
         ~40-60 words). This is the explanatory body text.

IMPORTANT: You MUST use exactly the labels "Insight: " and "Why it matters: " — 
these labels are required for formatting.

Example format:
Insight: Drug spans 4 therapy areas beyond its original metabolic indication.
Why it matters: Multi-therapy-area reach transforms the commercial model from a single-franchise asset to a platform molecule. Each new therapy area unlocks distinct prescriber networks, payer segments, and revenue pools — compounding lifecycle value.

Insight: 3 secondary indications are already in Phase 3, signaling near-term label expansion.
Why it matters: Phase 3 secondary indications with active enrollment represent 12-24 month catalysts for label expansion. Successful readouts would broaden the addressable market and strengthen payer negotiation leverage with real-world evidence of multi-indication utility.

Be specific. Reference indication counts, therapy area breadth, primary vs. secondary 
classification, geographic reach, or data source corroboration. Do NOT write generic 
statements like "the label is expanding" without concrete data.

Prioritize insights that address:
1. Breadth of indication portfolio and what it enables (platform potential, lifecycle value)
2. Therapy area diversification and cross-specialty commercial opportunity
3. Regulatory label status — approved indications vs. investigational expansions
4. Quality and variety of evidence sources (clinical trials, regulatory filings, investor disclosures)
5. Label expansion gaps that create risk or delay
6. Pipeline maturity of secondary indications (how close to approval)

## EXPANSION INDICATIONS
For EACH secondary (expansion) indication listed in the data below, write EXACTLY 
two lines using this format:
Line 1: "Indication: " followed by the indication name and its therapy area in parentheses.
Line 2: "Rationale: " followed by 1-2 sentences explaining WHY this is classified as 
         a secondary/expansion indication — i.e., what makes it distinct from the primary 
         label, what evidence supports it, and its current development status. Use the 
         rationale data provided but rewrite it in clear business language.

IMPORTANT: You MUST use exactly the labels "Indication: " and "Rationale: " — 
these labels are required for formatting. Cover ALL secondary indications listed.

## EVIDENCE GAPS & RISKS
Write 3-5 bullet points (each starting with "- ") identifying the most material 
gaps in the label expansion profile and the business risk each creates. Focus ONLY 
on indication and expansion gaps — for example, missing indications in large 
addressable markets, limited expansion beyond primary therapy area, over-reliance 
on a single indication for revenue, absence of real-world evidence for newer 
indications, or lack of confirmatory data for pipeline expansions.
CRITICAL: Do NOT mention peer-reviewed journals, publications, published literature, 
academic publishing, or the need for more published studies. These are NOT relevant 
gaps for this report. Every gap must be about missing DATA or missing INDICATIONS.

## BOTTOM LINE
Write 3-4 sentences stating what a decision-maker should infer from this dimension. 
Be direct and actionable — state whether the label expansion profile supports 
investment, partnership, or market entry decisions, and flag any conditions or 
watchpoints. Focus on the strategic value of the indication breadth.

STRICT RULES:
- Total length: 700-1000 words (the report should comfortably fill ~2 pages)
- NO technical jargon (no "Ep", "Et", "scoring", "model", "pipeline page API", 
  "BigQuery", "ClinicalTrials.gov API", "Gemini", "LLM")
- Do NOT mention scores of any kind — no Ep, Et, numerical scores, or scoring methodology
- Do NOT mention peer-reviewed journals, publications, or academic publishing anywhere
- You MUST use "Insight: " and "Why it matters: " labels exactly in KEY INSIGHTS
- Every statement must add insight or implication — no restating obvious facts
- Use clear, natural business language that a non-scientific executive can follow
- Do not use markdown bold (**text**) — use plain text only
- Reference specific numbers, indication counts, and therapy areas wherever possible
- Keep paragraphs short (2-4 sentences max)

DATA FOR YOUR ANALYSIS:
======================

Molecule: {molecule_name}
Company: {data['company_name'] or 'Unknown'}

Total Unique Indications: {data['total_indications']}
Primary Indications: {primary_list}
Secondary (Expansion) Indications: {secondary_list}
Number of Therapy Areas: {len(data['therapy_areas'])}
Therapy Areas: {ta_list}
Total Trials / Sources: {data['total_trials']}
Data Sources: {sources_list}
Has Regulatory Label Data: {'Yes' if data['has_regulatory_label'] else 'No'}

Phase Distribution: {json.dumps(data['phase_distribution'])}

Secondary Indication Details (with rationale for each):
{chr(10).join(secondary_rationale_lines) if secondary_rationale_lines else 'No secondary indications identified'}

Indication Details (first 30):
{chr(10).join(indication_lines) if indication_lines else 'No indication details available'}

Now write the report. Remember: business language, specific numbers, no jargon, 
no scores, 700-1000 words. 
CRITICAL: In KEY INSIGHTS, every insight headline MUST start with "Insight: " 
and every body line MUST start with "Why it matters: "."""

    return prompt


def generate_gemini_narrative(molecule_name: str, data: dict) -> str | None:
    import urllib.request
    import urllib.error

    api_key = os.environ.get("GEMINI_API_KEY", "").strip()
    if not api_key:
        print("  [WARN] GEMINI_API_KEY not set — skipping AI narrative.")
        return None

    prompt = build_gemini_prompt(molecule_name, data)

    payload = json.dumps({
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {
            "temperature": 0.25,
            "maxOutputTokens": 16384,
        },
    }).encode("utf-8")

    url = f"{GEMINI_API_URL}?key={api_key}"
    req = urllib.request.Request(
        url, data=payload,
        headers={"Content-Type": "application/json"},
        method="POST",
    )

    try:
        print(f"  [AI]  Requesting business narrative for {molecule_name}...")
        with urllib.request.urlopen(req, timeout=90) as resp:
            result = json.loads(resp.read().decode("utf-8"))
        text = (
            result.get("candidates", [{}])[0]
                  .get("content", {})
                  .get("parts", [{}])[0]
                  .get("text", "")
        )
        if text.strip():
            print(f"  [AI]  Narrative received ({len(text)} chars).")
            return text.strip()
        else:
            print("  [WARN] Gemini returned empty response.", file=sys.stderr)
            return None
    except Exception as e:
        print(f"  [ERR] Gemini API call failed: {e}", file=sys.stderr)
        return None


# ---------------------------------------------------------------------
# PDF building — compact, insight-driven, max ~2 pages
# ---------------------------------------------------------------------

NAVY = colors.HexColor("#1F3864")
BLUE = colors.HexColor("#2E75B6")
LIGHT_BLUE = colors.HexColor("#D9E1F2")
GREEN = colors.HexColor("#2E7D32")
RED = colors.HexColor("#C62828")
AMBER = colors.HexColor("#B8860B")
GREY = colors.HexColor("#666666")
DARK_TEXT = colors.HexColor("#1A1A2E")


def build_styles():
    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(
        name="ReportTitle", fontSize=18, leading=22, textColor=NAVY,
        fontName="Helvetica-Bold", spaceAfter=2, alignment=TA_LEFT,
    ))
    styles.add(ParagraphStyle(
        name="ReportSubtitle", fontSize=9.5, leading=12, textColor=GREY,
        fontName="Helvetica", spaceAfter=1,
    ))
    styles.add(ParagraphStyle(
        name="SectionHeader", fontSize=11.5, leading=14, textColor=colors.white,
        fontName="Helvetica-Bold", spaceBefore=10, spaceAfter=0,
        backColor=NAVY, leftIndent=6, borderPadding=(5, 5, 5, 5),
    ))
    styles.add(ParagraphStyle(
        name="Headline", fontSize=11, leading=14, textColor=NAVY,
        fontName="Helvetica-Bold", spaceAfter=6, spaceBefore=6,
        alignment=TA_LEFT,
    ))
    styles.add(ParagraphStyle(
        name="InsightHeadline",
        fontSize=10.5,
        leading=14,
        textColor=BLUE,
        fontName="Helvetica-Bold",
        spaceBefore=10,
        spaceAfter=2,
        leftIndent=0,
        alignment=TA_LEFT,
    ))
    styles.add(ParagraphStyle(
        name="InsightBody",
        fontSize=9,
        leading=13,
        textColor=DARK_TEXT,
        fontName="Helvetica",
        spaceBefore=0,
        spaceAfter=6,
        leftIndent=0,
        alignment=TA_JUSTIFY,
    ))
    styles.add(ParagraphStyle(
        name="BulletText", fontSize=9, leading=13, textColor=DARK_TEXT,
        fontName="Helvetica-Bold", spaceAfter=1, spaceBefore=6,
        leftIndent=12, bulletIndent=0, alignment=TA_LEFT,
    ))
    styles.add(ParagraphStyle(
        name="BodyProse", fontSize=9, leading=13, textColor=DARK_TEXT,
        fontName="Helvetica", spaceAfter=3, alignment=TA_JUSTIFY,
    ))
    styles.add(ParagraphStyle(
        name="BottomLine", fontSize=9, leading=13, textColor=DARK_TEXT,
        fontName="Helvetica-Bold", spaceAfter=3, spaceBefore=2,
    ))
    styles.add(ParagraphStyle(
        name="SnapshotLabel", fontSize=8, leading=10, textColor=GREY,
        fontName="Helvetica-Bold", spaceAfter=0,
    ))
    styles.add(ParagraphStyle(
        name="SnapshotValue", fontSize=9, leading=11, textColor=DARK_TEXT,
        fontName="Helvetica-Bold", spaceAfter=0,
    ))
    styles.add(ParagraphStyle(
        name="FooterText", fontSize=7, leading=9, textColor=GREY,
        fontName="Helvetica",
    ))
    return styles


def parse_narrative_to_flowables(narrative: str, styles) -> list:
    """
    Parse the Gemini narrative into ReportLab flowables.

    KEY INSIGHTS section:
      Lines starting with "Insight: "       → InsightHeadline style (bold blue)
      Lines starting with "Why it matters:" → InsightBody style (dark text)

    All other sections map headings → SectionHeader, body → BodyProse.
    """
    flow = []
    current_section = None

    for line in narrative.split("\n"):
        stripped = line.strip()
        if not stripped:
            continue

        # Strip markdown bold markers
        stripped = stripped.replace("**", "")

        # ── Section heading detection ────────────────────────────────
        if stripped.startswith("## HEADLINE"):
            current_section = "headline"
            continue
        elif stripped.startswith("## INDICATION LANDSCAPE"):
            current_section = "landscape"
            flow.append(Spacer(1, 4))
            flow.append(Paragraph("INDICATION LANDSCAPE", styles["SectionHeader"]))
            flow.append(Spacer(1, 4))
            continue
        elif stripped.startswith("## KEY INSIGHTS"):
            current_section = "insights"
            flow.append(Spacer(1, 4))
            flow.append(Paragraph("KEY INSIGHTS", styles["SectionHeader"]))
            flow.append(Spacer(1, 2))
            continue
        elif stripped.startswith("## EVIDENCE GAPS"):
            current_section = "gaps"
            flow.append(Spacer(1, 4))
            flow.append(Paragraph("EVIDENCE GAPS &amp; RISKS", styles["SectionHeader"]))
            flow.append(Spacer(1, 4))
            continue
        elif stripped.startswith("## EXPANSION INDICATIONS"):
            current_section = "expansion"
            flow.append(Spacer(1, 4))
            flow.append(Paragraph("EXPANSION INDICATIONS", styles["SectionHeader"]))
            flow.append(Spacer(1, 2))
            continue
        elif stripped.startswith("## BOTTOM LINE"):
            current_section = "bottom"
            flow.append(Spacer(1, 4))
            flow.append(Paragraph("BOTTOM LINE", styles["SectionHeader"]))
            flow.append(Spacer(1, 3))
            continue
        elif stripped.startswith("## "):
            heading = stripped[3:].strip()
            current_section = "other"
            flow.append(Spacer(1, 4))
            flow.append(Paragraph(escape_html(heading), styles["SectionHeader"]))
            flow.append(Spacer(1, 3))
            continue

        # ── Content rendering by section ────────────────────────────
        if current_section == "headline":
            flow.append(Paragraph(escape_html(stripped), styles["Headline"]))

        elif current_section == "landscape":
            flow.append(Paragraph(escape_html(stripped), styles["BodyProse"]))

        elif current_section == "insights":
            # Strip leading list markers if Gemini adds them
            if stripped.startswith("- ") or stripped.startswith("* "):
                stripped = stripped[2:].strip()

            # ── "Insight: <headline text>" ───────────────────────────
            insight_match = re.match(r'^[Ii]nsight\s*\d*\s*[:.\-]\s*(.*)', stripped)
            if insight_match:
                headline_text = insight_match.group(1).strip()
                html = f'<font color="#2E75B6"><b>{escape_html(headline_text)}</b></font>'
                flow.append(Paragraph(html, styles["InsightHeadline"]))
                continue

            # ── "Why it matters: <body text>" ───────────────────────
            why_match = re.match(
                r'^[Ww]hy\s+[Ii]t\s+[Mm]atters\s*[:.\-]\s*(.*)', stripped
            )
            if why_match:
                body_text = why_match.group(1).strip()
                flow.append(Paragraph(escape_html(body_text), styles["InsightBody"]))
                continue

            # Fallback: any other line in insights section → plain bullet
            if stripped:
                flow.append(Paragraph(
                    f"&bull;&nbsp;&nbsp;{escape_html(stripped)}",
                    styles["BulletText"]
                ))

        elif current_section == "gaps":
            if stripped.startswith("- ") or stripped.startswith("* "):
                text = stripped[2:].strip()
            else:
                text = stripped
            # Skip lines about peer-reviewed journals / publications
            lower = text.lower()
            if any(kw in lower for kw in ["peer-review", "peer review", "journal",
                                           "publication", "published", "publishing"]):
                continue
            if text:
                flow.append(Paragraph(
                    f"&bull;&nbsp;&nbsp;{escape_html(text)}",
                    styles["BulletText"]
                ))

        elif current_section == "expansion":
            # Strip leading list markers if Gemini adds them
            if stripped.startswith("- ") or stripped.startswith("* "):
                stripped = stripped[2:].strip()

            # ── "Indication: <name>" ─────────────────────────────────
            ind_match = re.match(r'^[Ii]ndication\s*\d*\s*[:.\-]\s*(.*)', stripped)
            if ind_match:
                ind_text = ind_match.group(1).strip()
                html = f'<font color="#2E75B6"><b>{escape_html(ind_text)}</b></font>'
                flow.append(Paragraph(html, styles["InsightHeadline"]))
                continue

            # ── "Rationale: <explanation>" ────────────────────────────
            rat_match = re.match(r'^[Rr]ationale\s*[:.\-]\s*(.*)', stripped)
            if rat_match:
                body_text = rat_match.group(1).strip()
                flow.append(Paragraph(escape_html(body_text), styles["InsightBody"]))
                continue

            # Fallback: any other line → plain body text
            if stripped:
                flow.append(Paragraph(escape_html(stripped), styles["BodyProse"]))

        elif current_section == "bottom":
            flow.append(Paragraph(escape_html(stripped), styles["BottomLine"]))

        else:
            flow.append(Paragraph(escape_html(stripped), styles["BodyProse"]))

    return flow


def generate_pdf_report(molecule_name: str, data: dict, out_path: Path,
                        gemini_narrative: str | None = None,
                        innovator_name: str | None = None):
    """Generate a compact, business-facing PDF (target: 2 pages max)."""
    styles = build_styles()
    doc = SimpleDocTemplate(
        str(out_path), pagesize=letter,
        topMargin=0.5 * inch, bottomMargin=0.5 * inch,
        leftMargin=0.6 * inch, rightMargin=0.6 * inch,
        title=f"Label Expansion Report - {molecule_name}",
    )
    story = []

    generic_name = data.get("molecule_name", molecule_name)
    sponsor = innovator_name or data.get("company_name", "Unknown")

    # ---------- Title ----------
    story.append(Paragraph("Label Expansion Report", styles["ReportTitle"]))
    story.append(Paragraph(
        f"Molecule: <b>{escape_html(generic_name)}</b>"
        f"&nbsp;&nbsp;|&nbsp;&nbsp;Innovator: {escape_html(sponsor)}"
        f"&nbsp;&nbsp;|&nbsp;&nbsp;{datetime.now().strftime('%B %d, %Y')}",
        styles["ReportSubtitle"]
    ))

    # Show primary indications and therapy areas in subtitle
    primary = data.get("primary_indications", [])
    therapy_areas = data.get("therapy_areas", [])
    if primary or therapy_areas:
        meta_parts = []
        if primary:
            meta_parts.append(f"Primary: {', '.join(primary[:4])}")
        if therapy_areas:
            meta_parts.append(f"Therapy Areas: {', '.join(therapy_areas[:5])}")
        story.append(Paragraph("&nbsp;&nbsp;|&nbsp;&nbsp;".join(meta_parts),
                               styles["ReportSubtitle"]))

    story.append(Spacer(1, 4))
    story.append(HRFlowable(width="100%", thickness=1, color=NAVY))
    story.append(Spacer(1, 6))

    # ---------- Snapshot bar (compact stats) ----------
    total_indications = data.get("total_indications", 0)
    n_primary = len(data.get("primary_indications", []))
    n_secondary = len(data.get("secondary_indications", []))
    n_therapy_areas = len(data.get("therapy_areas", []))

    snap_data = [
        ("Indications", str(total_indications)),
        ("Primary", str(n_primary)),
        ("Secondary", str(n_secondary)),
        ("Therapy Areas", str(n_therapy_areas)),
    ]
    snap_cells = []
    for label, val in snap_data:
        snap_cells.append([
            Paragraph(val, ParagraphStyle("sv", parent=styles["SnapshotValue"],
                                          textColor=NAVY)),
            Paragraph(label, styles["SnapshotLabel"]),
        ])

    snap_table = Table([snap_cells], colWidths=[1.65 * inch] * 4,
                       rowHeights=[0.4 * inch])
    snap_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F5F7FA")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E0E0E0")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(snap_table)
    story.append(Spacer(1, 8))

    # ---------- Gemini narrative (the core of the report) ----------
    if gemini_narrative:
        flow = parse_narrative_to_flowables(gemini_narrative, styles)
        story.extend(flow)
    else:
        # Fallback: structured summary from data
        story.append(Paragraph("EXECUTIVE SUMMARY", styles["SectionHeader"]))
        story.append(Spacer(1, 4))

        primary_str = ", ".join(data.get("primary_indications", [])[:5]) or "not identified"
        secondary_str = ", ".join(data.get("secondary_indications", [])[:5]) or "none identified"
        ta_str = ", ".join(data.get("therapy_areas", [])) or "not classified"

        summary = (
            f"{generic_name} has {total_indications} distinct indications identified. "
            f"Primary indications include {primary_str}. "
            f"Secondary (expansion) indications include {secondary_str}. "
            f"The molecule spans {n_therapy_areas} therapy area(s): {ta_str}."
        )
        story.append(Paragraph(escape_html(summary), styles["BodyProse"]))

        secondary_inds = data.get("secondary_indications", [])
        rationales_map = data.get("secondary_indication_rationales", {})
        if secondary_inds:
            story.append(Spacer(1, 4))
            story.append(Paragraph("EXPANSION INDICATIONS", styles["SectionHeader"]))
            story.append(Spacer(1, 3))
            for ind in secondary_inds[:10]:
                info = rationales_map.get(ind, {})
                rationale = info.get("rationale", "")
                ta = info.get("therapy_area", "")
                phase = info.get("phase", "")

                # Headline: indication name (bold blue)
                html = f'<font color="#2E75B6"><b>{escape_html(ind)}</b></font>'
                if ta:
                    html += f'&nbsp;&nbsp;<font color="#666666">({escape_html(ta)})</font>'
                if phase:
                    html += f'&nbsp;&nbsp;<font color="#666666">| {escape_html(phase)}</font>'
                story.append(Paragraph(html, styles["InsightHeadline"]))

                # Body: rationale explaining why it is secondary
                if rationale:
                    story.append(Paragraph(escape_html(rationale), styles["InsightBody"]))
                else:
                    story.append(Paragraph(
                        "Classified as a label expansion beyond the primary approved indication.",
                        styles["InsightBody"]
                    ))

        story.append(Spacer(1, 6))
        story.append(Paragraph(
            "AI narrative was not generated (GEMINI_API_KEY missing or API error).",
            styles["BodyProse"]
        ))

    # ---------- Footer ----------
    story.append(Spacer(1, 10))
    story.append(HRFlowable(width="100%", thickness=0.5,
                            color=colors.HexColor("#CCCCCC")))
    footer_parts = [f"Report generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
    if gemini_narrative:
        footer_parts.append("Analytical narrative generated by Gemini 2.5 Flash")
    story.append(Paragraph("  |  ".join(footer_parts), styles["FooterText"]))

    doc.build(story)


# ---------------------------------------------------------------------
# Per-molecule processing
# ---------------------------------------------------------------------

def process_molecule(molecule_name: str, output_dir: str, report_dir: Path) -> bool:
    excel_path = find_excel_file(output_dir, molecule_name)
    if not excel_path:
        print(f"  [ERR] No Excel file found for '{molecule_name}' in {output_dir}",
              file=sys.stderr)
        print(f"        Run run_all.py or label.py first to generate data.",
              file=sys.stderr)
        return False

    print(f"  [OK]  Found Excel file: {excel_path}")

    # Read and extract data
    rows = read_excel_data(excel_path)
    if not rows:
        print(f"  [ERR] No data rows found in {excel_path}", file=sys.stderr)
        return False

    data = extract_label_data(rows, molecule_name)
    resolved_name = data["molecule_name"]
    print(f"  [OK]  Extracted data for: {resolved_name}")
    print(f"        Indications: {data['total_indications']} | "
          f"Primary: {len(data['primary_indications'])} | "
          f"Secondary: {len(data['secondary_indications'])} | "
          f"Therapy Areas: {len(data['therapy_areas'])}")

    # Generate narrative
    gemini_narrative = generate_gemini_narrative(resolved_name, data)

    # Resolve innovator
    company_hint = data.get("company_name", "")
    innovator_name = lookup_innovator(resolved_name, company_hint) or company_hint or None

    safe = re.sub(r"[^a-zA-Z0-9]", "_", str(resolved_name).lower())
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = report_dir / f"{safe}_label_expansion_report_{ts}.pdf"

    generate_pdf_report(
        molecule_name=resolved_name,
        data=data,
        out_path=out_path,
        gemini_narrative=gemini_narrative,
        innovator_name=innovator_name,
    )
    print(f"  [OK]  PDF written -> {out_path}")
    return True


# ---------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Generate business-facing PDF label expansion reports from Excel data"
    )
    parser.add_argument(
        "--molecule", "-m", nargs="+", required=True, metavar="MOLECULE",
        help="One or more molecule names",
    )
    parser.add_argument(
        "--output-dir", default=DEFAULT_OUTPUT_DIR,
        help=f"Directory containing *_research.xlsx or *_clinical_efficacy.xlsx files "
             f"(default: '{DEFAULT_OUTPUT_DIR}')",
    )
    parser.add_argument(
        "--report-dir", default=None,
        help="Where to save the generated PDFs (default: same as --output-dir)",
    )
    args = parser.parse_args()

    output_dir = args.output_dir
    report_dir = Path(args.report_dir) if args.report_dir else Path(output_dir)
    report_dir.mkdir(parents=True, exist_ok=True)

    molecules = args.molecule
    print(f"\nProcessing {len(molecules)} molecule(s): {', '.join(molecules)}\n")

    results = {}
    for molecule in molecules:
        print(f"--- {molecule} ---")
        ok = process_molecule(molecule, output_dir, report_dir)
        results[molecule] = ok
        print()

    succeeded = [m for m, ok in results.items() if ok]
    failed = [m for m, ok in results.items() if not ok]

    print("=" * 50)
    print(f"Done. {len(succeeded)}/{len(molecules)} report(s) generated successfully.")
    if failed:
        print(f"Failed: {', '.join(failed)}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
