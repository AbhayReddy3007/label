"""
scoring.py — Targeted OpenTargets Scoring + Maturity Weight Scoring
=====================================================================

Standalone, re-runnable step. Given an already-generated label-expansion
workbook (produced once by the full label_expansion pipeline — Module A
and Module B need NOT be re-run), this script:

  1. Reads "Clinical Efficacy" and "Drug Indication Research" directly
     from the workbook.
  2. Resolves the drug's MOA(s) to Ensembl target ID(s) ONCE, at drug
     level — not per row. The "moa" field can contain several MOAs,
     separated by ";", and it's stamped identically onto every row for a
     single-drug workbook, so it only needs resolving once:
       a. Gemini + Search first finds the OpenTargets-equivalent target
          name for each MOA.
       b. The OpenTargets search API resolves that name to its real
          Ensembl ID (never trusting an LLM-generated ID directly).
  3. For every row whose indication_type is "Secondary":
       a. Resolves the row's indication to its OpenTargets-equivalent
          disease name — again in two steps: Gemini + Search first
          (handles abbreviations/synonyms), then the OpenTargets
          `search` API confirms the canonical name and returns its EFO
          ID. The indication cell is renamed to this canonical name.
       b. Queries the OpenTargets target-disease association score
          (`associatedDiseases(Bs: [efoId])`, filtered to that one
          disease) for every one of the drug's resolved Ensembl targets
          (from step 2), and keeps the HIGHEST score across them — this
          is the drug's best-supported target for that indication.
     This never bulk-fetches "all associated diseases" for a target —
     only the specific indications already present in our own data are
     looked up. Primary rows are left completely untouched: no
     OpenTargets or Gemini call is made for them, and any pre-existing
     opentargets_score on a non-Secondary row is cleared.
  4. Rebuilds "All Combined" from the (now updated) Clinical Efficacy +
     Drug Indication Research rows only — the old bulk "OpenTargets
     Indications" dump is no longer folded in, so this sheet only ever
     contains indications actually observed in trials / innovator
     research, not every disease OpenTargets happens to associate with
     the target.
  5. Rebuilds the "Maturity Weight Scoring" sheet from that new "All
     Combined" data — keeping Secondary rows ONLY (Primary rows are
     dropped from this sheet) — and appending Maturity_Weight (per row,
     from phase), Effective_Indications (drug-level — the sum of
     Maturity_Weight across all rows in the sheet, repeated on every
     row), and Prior (per row, from opentargets_score: >0.49→0.8,
     0.1-0.49→0.4, <0.1/unavailable→0.0).

Every step is idempotent — sheets are updated/replaced in place, not
appended to — and OpenTargets/Gemini lookups are cached in-memory per
run so the same MOA or indication is never looked up twice in one pass.

Usage (standalone)
-------------------
    python -m medical_potential.label_expansion.scoring path/to/workbook.xlsx

Programmatic usage
-------------------
    from medical_potential.label_expansion.scoring import run_scoring
    run_scoring("output/semaglutide_label_expansion.xlsx")
"""

from __future__ import annotations

import json
import logging
import re
import sys
import threading
import time
from pathlib import Path

import openpyxl
import requests
from google import genai
from google.genai import types as genai_types
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import medical_potential.config as config
from medical_potential.label_expansion.excel_writer import (
    HEADERS,
    COL_WIDTHS,
    _write_standard_sheet,
)

logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════
#  GEMINI CLIENT + PLUMBING
# ══════════════════════════════════════════════════════════════════════════
# Duplicated (trimmed) from research_modules.py rather than imported, so
# this file has no dependency on the rest of the pipeline and can be run
# completely on its own against an existing workbook.
_gemini_client: genai.Client | None = None
_gemini_client_lock = threading.Lock()


def _get_gemini_client() -> genai.Client:
    """Return a lazily-created, cached Gemini client (thread-safe)."""
    global _gemini_client
    if _gemini_client is None:
        with _gemini_client_lock:
            if _gemini_client is None:
                if not config.GEMINI_API_KEY:
                    raise RuntimeError(
                        "GEMINI_API_KEY is not set. Add it to your .env file or environment."
                    )
                _gemini_client = genai.Client(api_key=config.GEMINI_API_KEY)
    return _gemini_client


def _safe_response_text(resp) -> str:
    """Safely extract text from a Gemini response (see research_modules.py
    for the full explanation of why this walks the candidates/parts tree)."""
    try:
        if resp.text is not None:
            return resp.text.strip()
    except Exception:
        pass

    texts = []
    try:
        for candidate in (resp.candidates or []):
            try:
                parts = candidate.content.parts
            except Exception:
                continue
            for part in (parts or []):
                try:
                    t = getattr(part, "text", None)
                    if t:
                        texts.append(t.strip())
                except Exception:
                    continue
    except Exception:
        pass

    return "\n".join(texts) if texts else ""


def _is_transient_error(exc: Exception) -> bool:
    err = str(exc).lower()
    return any(k in err for k in (
        "503", "429", "unavailable", "overloaded", "resource exhausted",
        "rate limit", "deadline exceeded", "connection", "timeout", "502", "500",
    ))


def _gemini_generate(
    prompt: str,
    *,
    system_instruction: str = "",
    use_search: bool = True,
    model: str | None = None,
) -> str:
    """Call Gemini with optional Google Search grounding, with retries."""
    client = _get_gemini_client()
    model = model or config.GEMINI_FLASH_PREVIEW_MODEL
    max_retries = config.GEMINI_MAX_RETRIES
    base_delay = config.GEMINI_RETRY_BASE_DELAY_SECONDS

    configs = []
    if use_search:
        configs.append(genai_types.GenerateContentConfig(
            temperature=0,
            tools=[genai_types.Tool(google_search=genai_types.GoogleSearch())],
            system_instruction=system_instruction or "Return ONLY valid JSON.",
        ))
    configs.append(genai_types.GenerateContentConfig(
        temperature=0,
        system_instruction=system_instruction or "Return ONLY valid JSON.",
    ))

    for i, cfg in enumerate(configs):
        last_err = None
        for attempt in range(max_retries):
            try:
                resp = client.models.generate_content(model=model, contents=prompt, config=cfg)
                text = _safe_response_text(resp)
                text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text).strip()
                if text:
                    return text
                break
            except Exception as e:
                last_err = e
                if _is_transient_error(e) and attempt < max_retries - 1:
                    delay = base_delay * (2 ** attempt)
                    logger.warning("Gemini call failed (%s) — retrying in %ss (%d/%d)",
                                   e, delay, attempt + 1, max_retries)
                    time.sleep(delay)
                elif i == 0 and len(configs) > 1:
                    logger.info("Error with Search grounding (%s) — retrying without grounding", e)
                    break
                else:
                    raise
        else:
            if i == 0 and len(configs) > 1:
                logger.info("Retries exhausted with Search grounding — retrying without grounding")
                continue
            if last_err:
                raise last_err

        if i == 0 and len(configs) > 1:
            logger.info("Empty response with Search grounding — retrying without grounding")

    return ""


def _extract_json(text: str) -> dict | list:
    """Extract and parse the first JSON object or array from *text*."""
    if not text:
        raise ValueError("Cannot extract JSON from empty text")

    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    stripped = re.sub(r"^```(?:json)?\s*", "", text, flags=re.MULTILINE)
    stripped = re.sub(r"\s*```\s*$", "", stripped, flags=re.MULTILINE).strip()
    try:
        return json.loads(stripped)
    except json.JSONDecodeError:
        pass

    for open_ch, close_ch in (("{", "}"), ("[", "]")):
        start = stripped.find(open_ch)
        if start == -1:
            continue
        depth, end = 0, start
        in_str = False
        while end < len(stripped):
            ch = stripped[end]
            if ch == '"' and (end == 0 or stripped[end - 1] != '\\'):
                in_str = not in_str
            elif not in_str:
                if ch == open_ch:
                    depth += 1
                elif ch == close_ch:
                    depth -= 1
                    if depth == 0:
                        try:
                            return json.loads(stripped[start:end + 1])
                        except json.JSONDecodeError:
                            break
            end += 1

    raise ValueError(f"No valid JSON found in response (first 200 chars): {text[:200]}")


# ══════════════════════════════════════════════════════════════════════════
#  TARGETED OPENTARGETS LOOKUPS (no bulk "all diseases" fetch)
# ══════════════════════════════════════════════════════════════════════════
OPENTARGETS_GRAPHQL_URL = "https://api.platform.opentargets.org/api/v4/graphql"


def _resolve_target_name_via_gemini(moa: str) -> str | None:
    """Step 1 of MOA resolution: use Gemini + Search to find the
    equivalent target/gene name — as named on the OpenTargets Platform —
    for a given Mechanism of Action string. Returns a name only (e.g. an
    HGNC gene symbol like "GLP1R"), not an ID; the ID itself is resolved
    afterward via the OpenTargets search API so we never trust an
    LLM-generated Ensembl ID directly."""
    prompt = f"""
You are a bioinformatics expert.

Mechanism of Action (MOA): "{moa}"

From this MOA, identify the primary molecular target (protein / gene),
and give its name using the naming convention used on the OpenTargets
Platform (https://platform.opentargets.org/) — typically the HGNC gene
symbol (e.g. "GLP1R", "EGFR", "TNF").

Search the web to confirm the correct target name.

Return ONLY valid JSON — no markdown, no explanation:
{{"target_name": "<gene/protein symbol>"}}
If you cannot determine it, return:
{{"target_name": ""}}
"""
    try:
        text = _gemini_generate(prompt, system_instruction="Return ONLY valid JSON.", use_search=True)
        if not text:
            return None
        data = _extract_json(text)
        target_name = (data.get("target_name") or "").strip()
        return target_name or None
    except Exception as e:
        logger.warning("Failed to resolve OpenTargets target name for MOA '%s': %s", moa, e)
        return None


def _search_opentargets_target(name: str) -> tuple[str, str] | None:
    """Look up a target name against OpenTargets' search endpoint.
    Returns (ensembl_id, canonical_target_name) for the best hit, or None."""
    query_string = """
    query targetSearch($q: String!) {
      search(queryString: $q, entityNames: ["target"], page: {index: 0, size: 5}) {
        hits { id name entity }
      }
    }
    """
    try:
        resp = requests.post(
            OPENTARGETS_GRAPHQL_URL,
            json={"query": query_string, "variables": {"q": name}},
            timeout=20,
        )
        resp.raise_for_status()
        hits = resp.json().get("data", {}).get("search", {}).get("hits", []) or []
        for h in hits:
            if h.get("entity") == "target" and h.get("id") and h.get("name"):
                return h["id"], h["name"]
        logger.info("No OpenTargets target match for '%s'", name)
        return None
    except requests.RequestException as e:
        logger.warning("OpenTargets target search failed for '%s': %s", name, e)
        return None


def _resolve_ensembl_id_for_moa(moa: str) -> str | None:
    """Full MOA → Ensembl ID resolution, in two steps:
      1. Gemini + Search finds the OpenTargets-equivalent target name.
      2. The OpenTargets search API resolves that name to its real
         Ensembl gene ID (and canonical name), so we never rely on an
         LLM-generated ID directly.
    """
    target_name = _resolve_target_name_via_gemini(moa)
    if not target_name:
        logger.warning("Could not resolve an OpenTargets target name for MOA '%s'", moa)
        return None

    hit = _search_opentargets_target(target_name)
    if not hit:
        logger.warning("No OpenTargets target match for '%s' (from MOA '%s')", target_name, moa)
        return None

    ensembl_id, canonical_name = hit
    logger.info("MOA '%s' → target '%s' → Ensembl ID: %s", moa, canonical_name, ensembl_id)
    return ensembl_id


def _resolve_indication_name_via_gemini(indication: str) -> str | None:
    """Use Gemini + Search to find the equivalent disease name — as named
    on the OpenTargets Platform (EFO naming conventions) — for a given
    indication string. This runs BEFORE the OpenTargets disease search,
    so e.g. an abbreviation or a trial-specific phrasing gets normalized
    to something OpenTargets is more likely to recognize."""
    prompt = f"""
You are a medical terminology expert.

Indication: "{indication}"

Identify the equivalent disease/phenotype name as used on the OpenTargets
Platform (https://platform.opentargets.org/), which follows EFO
(Experimental Factor Ontology) naming conventions.

Search the web if needed to confirm the correct name.

Return ONLY valid JSON — no markdown, no explanation:
{{"disease_name": "<EFO-style disease name>"}}
If you cannot determine it, return:
{{"disease_name": ""}}
"""
    try:
        text = _gemini_generate(prompt, system_instruction="Return ONLY valid JSON.", use_search=True)
        if not text:
            return None
        data = _extract_json(text)
        disease_name = (data.get("disease_name") or "").strip()
        return disease_name or None
    except Exception as e:
        logger.warning("Failed to resolve OpenTargets disease name for indication '%s': %s", indication, e)
        return None


def _search_opentargets_disease(indication: str) -> tuple[str, str] | None:
    """Look up ONE specific indication by name against OpenTargets' search
    endpoint. Returns (efo_id, canonical_name) for the best disease hit,
    or None if nothing matched. This is a single targeted lookup — it does
    not fetch or page through OpenTargets' full disease list."""
    query_string = """
    query diseaseSearch($q: String!) {
      search(queryString: $q, entityNames: ["disease"], page: {index: 0, size: 5}) {
        hits { id name entity score }
      }
    }
    """
    try:
        resp = requests.post(
            OPENTARGETS_GRAPHQL_URL,
            json={"query": query_string, "variables": {"q": indication}},
            timeout=20,
        )
        resp.raise_for_status()
        hits = resp.json().get("data", {}).get("search", {}).get("hits", []) or []
        for h in hits:
            if h.get("entity") == "disease" and h.get("id") and h.get("name"):
                return h["id"], h["name"]
        logger.info("No OpenTargets disease match for indication '%s'", indication)
        return None
    except requests.RequestException as e:
        logger.warning("OpenTargets disease search failed for '%s': %s", indication, e)
        return None


def _get_target_disease_score(ensembl_id: str, efo_id: str) -> float | None:
    """Fetch the association score for exactly one (target, disease) pair.

    Uses the `Bs` filter on `target.associatedDiseases` to restrict the
    result to the single disease ID we care about, instead of paging
    through every disease associated with the target."""
    query_string = """
    query targetDiseaseScore($ensemblId: String!, $diseaseIds: [String!]) {
      target(ensemblId: $ensemblId) {
        associatedDiseases(Bs: $diseaseIds, page: {index: 0, size: 5}) {
          rows { score disease { id name } }
        }
      }
    }
    """
    variables = {"ensemblId": ensembl_id, "diseaseIds": [efo_id]}
    try:
        resp = requests.post(
            OPENTARGETS_GRAPHQL_URL,
            json={"query": query_string, "variables": variables},
            timeout=20,
        )
        resp.raise_for_status()
        target = resp.json().get("data", {}).get("target")
        if not target:
            return None
        for row in (target.get("associatedDiseases", {}).get("rows", []) or []):
            if row.get("disease", {}).get("id") == efo_id:
                return round(row.get("score", 0.0), 4)
        return None
    except requests.RequestException as e:
        logger.warning("OpenTargets target-disease lookup failed (%s, %s): %s", ensembl_id, efo_id, e)
        return None


# ══════════════════════════════════════════════════════════════════════════
#  SHEET READ / UPDATE — Clinical Efficacy & Drug Indication Research
# ══════════════════════════════════════════════════════════════════════════
def _read_sheet_rows(ws) -> tuple[list[str], list[dict], dict[str, int]]:
    """Read a standard-layout sheet (title row 1, metadata row 2, headers
    row 3, data from row 4). Returns (headers, row_dicts, header_to_col)."""
    header_row, data_start = 3, 4
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        headers.append(str(val) if val is not None else f"col_{col}")
    header_to_col = {h: i + 1 for i, h in enumerate(headers)}

    rows = []
    row_numbers = []
    for r in range(data_start, ws.max_row + 1):
        row_data = {h: ws.cell(row=r, column=header_to_col[h]).value for h in headers}
        if any(v is not None and str(v).strip() for v in row_data.values()):
            rows.append(row_data)
            row_numbers.append(r)

    for row, r in zip(rows, row_numbers):
        row["_row_number"] = r
    return headers, rows, header_to_col


def _resolve_drug_ensembl_ids(moa_field: str, ensembl_cache: dict[str, str | None]) -> list[str]:
    """Resolve every MOA listed for the drug (the "moa" field can contain
    several, separated by ";") to its Ensembl target ID — ONCE, at drug
    level. MOA is stamped identically onto every row of a single-drug
    workbook, so there is no need to re-resolve it per row. Returns a
    deduped list of Ensembl IDs (empty if none resolved)."""
    ensembl_ids: list[str] = []
    for moa in [m.strip() for m in moa_field.split(";") if m.strip()]:
        if moa not in ensembl_cache:
            ensembl_cache[moa] = _resolve_ensembl_id_for_moa(moa)
        eid = ensembl_cache[moa]
        if eid and eid not in ensembl_ids:
            ensembl_ids.append(eid)
    return ensembl_ids


def _get_drug_moa_field(wb) -> str:
    """Return the drug's "moa" field, read from the first row that has
    one across "Clinical Efficacy" and "Drug Indication Research". MOA is
    a drug-level property (stamped identically onto every row by
    research_modules.py), so any single row's value represents the whole
    drug — there's no need to scan every row."""
    for sheet_name in ("Clinical Efficacy", "Drug Indication Research"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        _headers, rows, _header_to_col = _read_sheet_rows(ws)
        for row in rows:
            moa = str(row.get("moa", "") or "").strip()
            if moa:
                return moa
    return ""


def _update_secondary_rows_with_opentargets(
    ws,
    header_to_col: dict[str, int],
    rows: list[dict],
    ensembl_ids: list[str],
    disease_cache: dict[str, tuple[str, str] | None],
) -> None:
    """Mutate *ws* in place: for Secondary rows, resolve the indication's
    OpenTargets-equivalent name (Gemini + Search, then confirmed via the
    OpenTargets search API), align the indication name, and stamp
    opentargets_score with the highest association score across the
    drug's *ensembl_ids* (resolved once, at drug level, by the caller —
    see _resolve_drug_ensembl_ids). For every other row, clear any stale
    opentargets_score. Cell writes are targeted (only the indication and
    opentargets_score columns) so all other formatting is untouched."""
    ind_col = header_to_col.get("indication")
    itype_col = header_to_col.get("indication_type")
    score_col = header_to_col.get("opentargets_score")

    if not (ind_col and itype_col and score_col):
        logger.warning("Sheet is missing indication/indication_type/opentargets_score columns — skipping")
        return

    for row in rows:
        r = row["_row_number"]
        itype = str(row.get("indication_type", "") or "").strip().lower()

        if itype != "secondary":
            # Not a Secondary row — make sure no stale score lingers here.
            if row.get("opentargets_score") not in ("", None):
                ws.cell(row=r, column=score_col, value="")
                row["opentargets_score"] = ""
            continue

        indication = str(row.get("indication", "") or "").strip()
        if not indication or indication.lower() == "no indication found":
            continue

        if not ensembl_ids:
            logger.info(
                "Secondary row '%s' has no resolvable OpenTargets target for this drug — skipping",
                indication,
            )
            continue

        # Resolve the indication's OpenTargets-equivalent name via
        # Gemini + Search first, then confirm/locate it (and its EFO ID)
        # through the OpenTargets search API.
        cache_key = indication.lower()
        if cache_key not in disease_cache:
            gemini_name = _resolve_indication_name_via_gemini(indication)
            disease_cache[cache_key] = _search_opentargets_disease(gemini_name or indication)
        hit = disease_cache[cache_key]
        if not hit:
            continue
        efo_id, ot_name = hit

        # Query the association score against every resolved drug-level
        # target and keep the highest one — this is the drug's
        # best-supported target for this indication.
        best_score = None
        for ensembl_id in ensembl_ids:
            score = _get_target_disease_score(ensembl_id, efo_id)
            if score is not None and (best_score is None or score > best_score):
                best_score = score
        if best_score is None:
            continue
        score = best_score

        # Align the indication name to OpenTargets' canonical name.
        ws.cell(row=r, column=ind_col, value=ot_name)
        row["indication"] = ot_name

        score_cell = ws.cell(row=r, column=score_col, value=score)
        score_cell.number_format = '0.0000'
        row["opentargets_score"] = score


def _process_sheet(wb, sheet_name: str, ensembl_ids: list[str], disease_cache: dict) -> list[dict]:
    """Read *sheet_name*, apply targeted OpenTargets scoring to its
    Secondary rows in place (using the drug-level *ensembl_ids* resolved
    once by the caller), and return the resulting row dicts (ready for
    use elsewhere, e.g. rebuilding "All Combined")."""
    if sheet_name not in wb.sheetnames:
        logger.warning("Sheet '%s' not found in workbook — skipping", sheet_name)
        return []
    ws = wb[sheet_name]
    _headers, rows, header_to_col = _read_sheet_rows(ws)
    _update_secondary_rows_with_opentargets(ws, header_to_col, rows, ensembl_ids, disease_cache)
    for row in rows:
        row.pop("_row_number", None)
    return rows


# ══════════════════════════════════════════════════════════════════════════
#  PHASE → MATURITY WEIGHT MAPPING
# ══════════════════════════════════════════════════════════════════════════
def _phase_to_maturity_weight(phase) -> float:
    """Map a phase string to its maturity weight.

    Rules:
      - "Not available", "Preclinical", empty, or unrecognised → 0.05
      - "Phase 1" (or contains "1")                           → 0.1
      - "Phase 2" (or contains "2")                           → 0.3
      - "Phase 3" (or contains "3")                           → 0.6
      - "Approved", "Phase 4", "Marketed", "Launched"         → 1.0
    """
    raw = str(phase).strip().lower() if phase else ""

    if not raw or raw in ("not available", "preclinical", "n/a", "none", ""):
        return 0.05

    if any(kw in raw for kw in ("approved", "marketed", "launched")):
        return 1.0

    nums = re.findall(r"\d+", raw)
    if nums:
        max_phase = max(int(n) for n in nums)
        if max_phase >= 4:
            return 1.0
        elif max_phase == 3:
            return 0.6
        elif max_phase == 2:
            return 0.3
        elif max_phase == 1:
            return 0.1

    if "preclinical" in raw or "pre-clinical" in raw:
        return 0.05

    return 0.05


# ══════════════════════════════════════════════════════════════════════════
#  OPENTARGETS SCORE → PRIOR MAPPING
# ══════════════════════════════════════════════════════════════════════════
def _opentargets_score_to_prior(score) -> float:
    """Map an opentargets_score value to its prior.

    Rules:
      - score > 0.49            → 0.8
      - 0.1 <= score <= 0.49    → 0.4
      - score < 0.1 or unavailable → 0.0
    """
    try:
        value = float(score)
    except (TypeError, ValueError):
        return 0.0

    if value > 0.49:
        return 0.8
    elif value >= 0.1:
        return 0.4
    else:
        return 0.0


# ══════════════════════════════════════════════════════════════════════════
#  ALL COMBINED — rebuilt from Clinical Efficacy + Drug Indication
#  Research only (the bulk "OpenTargets Indications" dump is NOT folded
#  back in, so this sheet only ever reflects indications actually
#  observed in trials / innovator research).
# ══════════════════════════════════════════════════════════════════════════
def _rebuild_all_combined(wb, clinical_rows: list[dict], indication_rows: list[dict], molecule: str) -> None:
    combined = clinical_rows + indication_rows

    et_values = [r.get("Et") for r in combined if r.get("Et")]
    max_et = ""
    if et_values:
        try:
            max_et = str(max(int(v) for v in et_values if str(v).isdigit()))
        except ValueError:
            max_et = et_values[0]
    for row in combined:
        row["Et"] = max_et

    if "All Combined" in wb.sheetnames:
        idx = wb.sheetnames.index("All Combined")
        del wb["All Combined"]
    else:
        idx = len(wb.sheetnames)

    ws = wb.create_sheet("All Combined")
    _write_standard_sheet(ws, combined, molecule, "All Sources Combined", headers=HEADERS, col_widths=COL_WIDTHS)
    wb.move_sheet("All Combined", offset=idx - (len(wb.sheetnames) - 1))


# ══════════════════════════════════════════════════════════════════════════
#  MATURITY WEIGHT SHEET WRITER
# ══════════════════════════════════════════════════════════════════════════
def add_maturity_weight_sheet(excel_path: str | Path) -> Path:
    """Open the label-expansion workbook, read the "All Combined" sheet,
    keep Secondary-indication rows only, and add a new "Maturity Weight
    Scoring" sheet with three columns appended:

      - Maturity_Weight: per-row, derived from that row's phase.
      - Effective_Indications: drug-level (same value on every row) —
        the sum of Maturity_Weight across all rows in this sheet (all of
        which are Secondary, since Primary rows are dropped).
      - Prior: per-row, derived from that row's opentargets_score
        (score > 0.49 → 0.8, 0.1–0.49 → 0.4, < 0.1 or unavailable → 0.0).

    Parameters
    ----------
    excel_path : str or Path
        Path to the existing .xlsx workbook.

    Returns
    -------
    Path to the (updated) workbook.
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Workbook not found: {excel_path}")

    wb = openpyxl.load_workbook(str(excel_path))

    source_sheet_name = "All Combined"
    if source_sheet_name not in wb.sheetnames:
        logger.warning("'%s' sheet not found in workbook — cannot add maturity weights", source_sheet_name)
        return excel_path

    src = wb[source_sheet_name]

    header_row = 3
    data_start = 4

    headers = []
    for col in range(1, src.max_column + 1):
        val = src.cell(row=header_row, column=col).value
        headers.append(str(val) if val is not None else f"col_{col}")

    data_rows = []
    for row_idx in range(data_start, src.max_row + 1):
        row_data = {}
        for col_idx, header in enumerate(headers, 1):
            row_data[header] = src.cell(row=row_idx, column=col_idx).value
        if any(v is not None and str(v).strip() for v in row_data.values()):
            data_rows.append(row_data)

    if not data_rows:
        logger.warning("No data rows found in '%s' — skipping maturity weight sheet", source_sheet_name)
        return excel_path

    # ── Keep Secondary rows only. The Maturity Weight Scoring sheet is
    # scoped to Secondary indications — Primary rows are dropped here,
    # before any of the downstream weight/score calculations run. ───────
    indication_type_col = None
    for h in headers:
        if h.lower() == "indication_type":
            indication_type_col = h
            break

    if indication_type_col:
        before = len(data_rows)
        data_rows = [
            r for r in data_rows
            if str(r.get(indication_type_col, "") or "").strip().lower() == "secondary"
        ]
        logger.info(
            "Maturity Weight Scoring: kept %d Secondary row(s) out of %d",
            len(data_rows), before,
        )
    else:
        logger.warning(
            "No 'indication_type' column found in '%s' — cannot filter to Secondary rows; "
            "including all rows", source_sheet_name,
        )

    if not data_rows:
        logger.warning("No Secondary rows found — skipping maturity weight sheet")
        return excel_path

    phase_col = None
    for h in headers:
        if h.lower() == "phase":
            phase_col = h
            break
    if not phase_col:
        logger.warning("No 'phase' column found in '%s' — cannot compute maturity weights", source_sheet_name)
        return excel_path

    for row in data_rows:
        phase_val = row.get(phase_col, "")
        row["Maturity_Weight"] = _phase_to_maturity_weight(phase_val)

    # ── Effective Indications: sum of Maturity_Weight. Every row here is
    # already Secondary (Primary rows were dropped above), so this is
    # simply the sum across all remaining rows. This is a drug-level
    # figure (like Et), so the same value is repeated on every row rather
    # than being row-specific. ───────────────────────────────────────────
    effective_indications = round(sum(row["Maturity_Weight"] for row in data_rows), 2)
    for row in data_rows:
        row["Effective_Indications"] = effective_indications

    # ── Prior: per-row, derived from that row's opentargets_score. Since
    # opentargets_score is only ever populated on Secondary rows (see the
    # defensive cleanup above), Primary rows naturally fall into the
    # "unavailable" → 0.0 case. ──────────────────────────────────────────
    for row in data_rows:
        row["Prior"] = _opentargets_score_to_prior(row.get("opentargets_score"))

    new_headers = headers + ["Maturity_Weight", "Effective_Indications", "Prior"]

    target_sheet_name = "Maturity Weight Scoring"
    if target_sheet_name in wb.sheetnames:
        del wb[target_sheet_name]

    ws = wb.create_sheet(target_sheet_name)

    thin = Border(
        bottom=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
    )
    hfill = PatternFill("solid", start_color="1A1A2E")
    alt_fill = PatternFill("solid", start_color="F4F7FB")
    ncols = len(new_headers)

    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value = "Maturity Weight Scoring"
    c.font = Font(name="Arial", bold=True, size=14, color="1A1A2E")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    c = ws["A2"]
    c.value = (
        f"Secondary indications only ({len(data_rows)} rows)  |  Weight mapping: Preclinical=0.05, "
        f"Ph1=0.1, Ph2=0.3, Ph3=0.6, Approved=1.0  |  "
        f"Effective_Indications = sum of Maturity_Weight across all rows in this sheet (drug-level)  |  "
        f"Prior: score>0.49→0.8, 0.1-0.49→0.4, <0.1/unavailable→0.0"
    )
    c.font = Font(name="Arial", italic=True, size=9, color="888888")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    for col, h in enumerate(new_headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin
    ws.row_dimensions[3].height = 22

    for idx, row in enumerate(data_rows, start=4):
        fill = alt_fill if idx % 2 == 0 else None
        for col, key in enumerate(new_headers, 1):
            val = row.get(key, "")
            c = ws.cell(row=idx, column=col, value=val)
            c.font = Font(name="Arial", size=9)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border = thin
            if fill:
                c.fill = fill

            if key == "Maturity_Weight" and isinstance(val, (int, float)):
                c.number_format = '0.00'
                c.alignment = Alignment(horizontal="center", vertical="top")

            if key == "Effective_Indications" and isinstance(val, (int, float)):
                c.number_format = '0.00'
                c.alignment = Alignment(horizontal="center", vertical="top")

            if key == "Prior" and isinstance(val, (int, float)):
                c.number_format = '0.00'
                c.alignment = Alignment(horizontal="center", vertical="top")

            if key == "opentargets_score" and isinstance(val, (int, float)):
                c.number_format = '0.0000'

        ws.row_dimensions[idx].height = 18

    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{3 + len(data_rows)}"
    ws.freeze_panes = "A4"

    standard_widths = {
        "molecule_name": 18, "company_name": 22, "indication": 28,
        "rationale": 40, "indication_type": 16, "therapy_area": 18,
        "TA-I": 36, "trial_title": 50, "trial_id": 18, "trial_size": 12,
        "phase": 10, "source_url": 40, "data_source": 16, "Ep": 8, "Et": 8,
        "moa": 28, "opentargets_score": 18, "Maturity_Weight": 16,
        "Effective_Indications": 20, "Prior": 12,
    }
    for i, h in enumerate(new_headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = standard_widths.get(h, 16)

    wb.save(str(excel_path))
    logger.info("Maturity Weight Scoring sheet added → %s", excel_path)
    return excel_path


# ══════════════════════════════════════════════════════════════════════════
#  ORCHESTRATION — the standalone, re-runnable entry point
# ══════════════════════════════════════════════════════════════════════════
def run_scoring(excel_path: str | Path) -> Path:
    """Full standalone scoring step. Run this on its own, any time, on an
    already-generated label-expansion workbook — it never touches
    BigQuery, GCS, or re-runs Module A / Module B.

    1. Resolve the drug's MOA(s) to Ensembl target ID(s) ONCE, at drug
       level (MOA is the same for every row of a single-drug workbook).
    2. Apply targeted OpenTargets scoring to Secondary rows in "Clinical
       Efficacy" and "Drug Indication Research" (in place), using that
       one drug-level resolution.
    3. Rebuild "All Combined" from just those two sheets.
    4. Rebuild "Maturity Weight Scoring" from the new "All Combined".
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Workbook not found: {excel_path}")

    wb = openpyxl.load_workbook(str(excel_path))

    ensembl_cache: dict[str, str | None] = {}
    disease_cache: dict[str, tuple[str, str] | None] = {}

    # ── Drug-level MOA resolution — happens ONCE, before any row is
    # processed, not per row. ────────────────────────────────────────────
    moa_field = _get_drug_moa_field(wb)
    if moa_field:
        drug_ensembl_ids = _resolve_drug_ensembl_ids(moa_field, ensembl_cache)
        if drug_ensembl_ids:
            logger.info(
                "Drug-level MOA resolution: '%s' → %d Ensembl target(s): %s",
                moa_field, len(drug_ensembl_ids), drug_ensembl_ids,
            )
        else:
            logger.warning(
                "Could not resolve any Ensembl target for MOA(s) '%s' — "
                "Secondary rows will not get an OpenTargets score", moa_field,
            )
    else:
        drug_ensembl_ids = []
        logger.warning(
            "No MOA found on any row — Secondary rows will not get an OpenTargets score"
        )

    clinical_rows = _process_sheet(wb, "Clinical Efficacy", drug_ensembl_ids, disease_cache)
    indication_rows = _process_sheet(wb, "Drug Indication Research", drug_ensembl_ids, disease_cache)

    molecule = ""
    for row in clinical_rows + indication_rows:
        if row.get("molecule_name"):
            molecule = str(row["molecule_name"])
            break

    _rebuild_all_combined(wb, clinical_rows, indication_rows, molecule)
    wb.save(str(excel_path))
    logger.info(
        "Targeted OpenTargets scoring applied and 'All Combined' rebuilt (%d clinical + %d indication rows) → %s",
        len(clinical_rows), len(indication_rows), excel_path,
    )

    return add_maturity_weight_sheet(excel_path)


# ══════════════════════════════════════════════════════════════════════════
#  CLI
# ══════════════════════════════════════════════════════════════════════════
def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)-8s | %(name)s | %(message)s",
    )
    if len(sys.argv) < 2:
        print("Usage: python -m medical_potential.label_expansion.scoring <path_to_xlsx>")
        sys.exit(1)

    path = sys.argv[1]
    run_scoring(path)


if __name__ == "__main__":
    main()
