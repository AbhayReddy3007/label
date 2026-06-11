"""
Unified Drug Research Pipeline
===============================
Runs all two modules and writes every result into a single Excel file,
one sheet per module, plus a combined sheet.

Modules:
  [1/2] Clinical Efficacy    — BigQuery clinical trial data enriched via Gemini
  [2/2] Drug Indication Research — Innovator sources (FDA labels, investor
         presentations, press releases, pipeline pages, SEC filings) via
         Gemini + Google Search

All sheets use the SAME standardized column format:
    molecule_name | company_name | indication | rationale | indication_type |
    therapy_area  | trial_title  | trial_id   | phase     | source_url      |
    data_source   | Ep           | Et

Indications are classified as Primary or Secondary by the LLM.
Therapy area is assigned by the LLM.
Each row has exactly ONE indication (multi-indication trials are unnested).

Usage:
    # Single molecule
    python run_all.py --molecule semaglutide
    python run_all.py --molecule semaglutide --company "Novo Nordisk"

    # Multiple selected molecules
    python run_all.py --drugs Semaglutide Liraglutide Tirzepatide

    # All molecules in BigQuery (one Excel file per molecule + a master file)
    python run_all.py --all
    python run_all.py --all --skip-existing   # skip molecules already processed
"""

import sys
import os
import re
import json
import time
import argparse
import requests
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from datetime import datetime

from dotenv import load_dotenv
load_dotenv(override=True)

from indication_standardizer import standardize_indication

# ── shared env ────────────────────────────────────────────────────────────────
GEMINI_API_KEY  = os.getenv("GEMINI_API_KEY", "").strip()
GBQ_PROJECT     = os.getenv("GBQ_PROJECT")
GBQ_DATASET     = os.getenv("GBQ_DATASET")
GBQ_TABLE       = os.getenv("GBQ_TABLE", "clinical_efficacy")
GBQ_SERVICE_KEY = os.getenv(
    "GBQ_SERVICE_KEY",
    r"C:\Users\p90022569\Downloads\cognito-prod-394707-d38a0283cb16 (2).json",
)

GEMINI_API_URL = (
    "https://generativelanguage.googleapis.com/v1beta/models/"
    "gemini-2.5-flash:generateContent"
)

# ── Parallel config ───────────────────────────────────────────────────────────
MAX_WORKERS_TRIALS    = 10   # parallel Gemini calls for clinical trial enrichment
MAX_WORKERS_INNOVATOR = 5    # parallel Gemini calls for innovator source queries

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════════════
#  SECONDARY INDICATION CRITERIA (from drug_indication_researcher.py)
# ══════════════════════════════════════════════════════════════════════════════

SECONDARY_INDICATION_CRITERIA = """
A secondary indication qualifies ONLY if ALL of the following are true:

- The indication represents a true expansion — i.e., it is not part of the primary indication (for clinical assets) or currently approved label (for commercial assets)
- The indication is described at a clear disease-level definition, avoiding vague, overlapping, or synonymous representations
- The source must describe observed or measured outcomes in that specific indication (e.g., trial results, endpoint readouts, biomarker response), not just planned evaluation or exploratory intent

The following must NOT be considered secondary indications:
* Indications mentioned only as hypothesis, targets, or exploratory possibilities
* Pipeline indications without any data or outcomes
* Mechanism based assumptions without clinical or empirical validation
* Early discovery or preclinical signals without human data
* Any indication lacking traceable, verifiable evidence of results
"""


# ══════════════════════════════════════════════════════════════════════════════
#  COMMON FORMAT — all sheets share these columns
# ══════════════════════════════════════════════════════════════════════════════

HEADERS    = ["molecule_name", "company_name", "indication", "rationale",
              "indication_type", "therapy_area", "trial_title", "trial_id",
              "phase", "source_url", "data_source", "Ep", "Et"]
COL_WIDTHS = [18, 22, 28, 40, 16, 18, 50, 18, 10, 40, 16, 8, 8]


# ══════════════════════════════════════════════════════════════════════════════
#  Ep SCORE — row-wise, based on phase
# ══════════════════════════════════════════════════════════════════════════════

def compute_ep(phase) -> str:
    """
    Ep score based on clinical trial phase (row-wise).
    Phase 4 or 3 → 5
    Phase 2     → 4
    Phase 1     → 3
    """
    phase_str = str(phase).strip().lower() if phase else ""
    nums = re.findall(r'\d+', phase_str)
    if not nums:
        if any(kw in phase_str for kw in ("approved", "launched", "marketed")):
            return "5"
        return ""
    max_phase = max(int(n) for n in nums)
    if max_phase >= 3:
        return "5"
    elif max_phase == 2:
        return "4"
    elif max_phase == 1:
        return "3"
    return ""


# ══════════════════════════════════════════════════════════════════════════════
#  Et SCORE — drug-level, based on therapy area & indication breadth
# ══════════════════════════════════════════════════════════════════════════════

def compute_et(enriched_rows, molecule_name, gemini_client=None, gen_types=None):
    """
    Et score computed across all rows for a drug:
      Multiple therapy areas                              → 5
      1 therapy area, 2+ secondary indications            → 4
      1 therapy area, 1 secondary (broad)                 → 3
      1 therapy area, 1 secondary (very niche)            → 2
      1 therapy area, primary indication(s) only          → 1
    """
    drug_rows = [r for r in enriched_rows
                 if r.get("molecule_name", "").lower() == molecule_name.lower()]
    if not drug_rows:
        return ""

    therapy_areas = set(
        r["therapy_area"] for r in drug_rows
        if r.get("therapy_area") and r["therapy_area"].lower() not in ("other", "")
    )
    secondary_indications = list(set(
        r["indication"] for r in drug_rows
        if r.get("indication_type", "").lower() == "secondary"
    ))
    primary_indications = list(set(
        r["indication"] for r in drug_rows
        if r.get("indication_type", "").lower() == "primary"
    ))

    if len(therapy_areas) > 1:
        return "5"
    elif len(secondary_indications) >= 2:
        return "4"
    elif len(secondary_indications) == 1:
        is_niche = _check_niche(molecule_name, primary_indications,
                                secondary_indications[0],
                                gemini_client, gen_types)
        return "2" if is_niche else "3"
    else:
        return "1"


def _check_niche(molecule, primary_list, secondary_indication,
                 gemini_client=None, gen_types=None):
    """Use Gemini to determine if a single secondary indication is very niche."""
    primary_str = ", ".join(primary_list) if primary_list else "unknown"
    prompt = f"""
You are a pharmaceutical analyst.

Drug: {molecule}
Primary indication(s): {primary_str}
Secondary (label expansion) indication: {secondary_indication}

Is this label expansion "very niche"?
A label expansion is "very niche" if it targets a narrow, specialized patient
sub-population, a rare disease, an orphan indication, or a highly specific
clinical context that affects relatively few patients compared to the primary
indication.

Return ONLY valid JSON:
{{"is_niche": true}} or {{"is_niche": false}}
No explanation.
"""
    try:
        if gemini_client is None:
            from google import genai as _genai
            from google.genai import types as _types
            gemini_client = _genai.Client(api_key=GEMINI_API_KEY)
            gen_types = _types

        resp = gemini_client.models.generate_content(
            model="gemini-2.5-flash",
            contents=prompt,
            config=gen_types.GenerateContentConfig(
                temperature=0,
                system_instruction="Return ONLY valid JSON.",
            ),
        )
        text = resp.text.strip()
        text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text).strip()
        data = json.loads(text)
        return data.get("is_niche", False)
    except Exception as e:
        print(f"   ⚠️  Niche check failed: {e} — defaulting to non-niche")
        return False


# ══════════════════════════════════════════════════════════════════════════════
#  BIGQUERY HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _bq_client():
    from google.cloud import bigquery
    from google.oauth2 import service_account
    credentials = service_account.Credentials.from_service_account_file(
        GBQ_SERVICE_KEY,
        scopes=["https://www.googleapis.com/auth/cloud-platform"],
    )
    return bigquery.Client(project=GBQ_PROJECT, credentials=credentials)


def fetch_all_molecules() -> list[str]:
    """Return the list of distinct molecule names in the BigQuery table."""
    print("  🔹 Fetching all distinct molecules from BigQuery…")
    client = _bq_client()
    query = f"""
        SELECT DISTINCT molecule_name
        FROM `{GBQ_PROJECT}.{GBQ_DATASET}.{GBQ_TABLE}`
        ORDER BY molecule_name
    """
    results = client.query(query).result()
    molecules = [row["molecule_name"] for row in results if row["molecule_name"]]
    print(f"  ✅ Found {len(molecules)} distinct molecules: {molecules}\n")
    return molecules


def fetch_bq_rows(molecule_name: str) -> list[dict]:
    print("  🔹 Connecting to BigQuery…")
    from google.cloud import bigquery
    client = _bq_client()
    query = f"""
        SELECT molecule_name, company_name, source_url, phase, trial_id
        FROM `{GBQ_PROJECT}.{GBQ_DATASET}.{GBQ_TABLE}`
        WHERE LOWER(molecule_name) = LOWER(@molecule)
    """
    job_config = bigquery.QueryJobConfig(
        query_parameters=[bigquery.ScalarQueryParameter("molecule", "STRING", molecule_name)]
    )
    results = client.query(query, job_config=job_config).result()
    rows = [dict(row) for row in results]
    print(f"  ✅ Retrieved {len(rows)} rows")
    return rows


# ══════════════════════════════════════════════════════════════════════════════
#  CHECKPOINT HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def load_checkpoint(path: str) -> dict:
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return {}

def save_checkpoint(path: str, data: dict):
    with open(path, "w") as f:
        json.dump(data, f, indent=2)


# ══════════════════════════════════════════════════════════════════════════════
#  GEMINI ENRICHMENT (clinical trials) — LLM-based classification
# ══════════════════════════════════════════════════════════════════════════════

def _enrich_single_trial(row, molecule_name, client, gen_types, checkpoint, cp_file, cp_lock):
    """Enrich a single trial. Returns (trial_id, conditions, trial_title, row)."""
    trial_id = row.get("trial_id")

    with cp_lock:
        if trial_id in checkpoint:
            data = checkpoint[trial_id]
            return (trial_id, data["conditions"], data["trial_title"], row)

    prompt = f"""
You are a clinical trial data assistant and pharmaceutical analyst.

Trial details:
Molecule: {row.get('molecule_name')}
Company:  {row.get('company_name')}
Trial ID: {trial_id}
Phase:    {row.get('phase')}
Source URL: {row.get('source_url')}

Your task:
1. Extract ALL disease indications being studied in this clinical trial.
2. Classify each indication as "Primary" or "Secondary".
3. Assign a therapy_area to each indication.

=== CLASSIFICATION RULES ===

"Primary" indication:
- The drug's main approved or originally intended indication
- The core disease the drug was designed and initially developed to treat
- What appears on the original FDA/EMA approved label

"Secondary" indication — must meet ALL of the following criteria:
{SECONDARY_INDICATION_CRITERIA}

=== THERAPY AREA ===
Assign one of: Metabolic, Cardiovascular, Oncology, Neuroscience, Immunology,
Respiratory, Nephrology, Hepatology, Ophthalmology, Musculoskeletal,
Gastroenterology, Infectious Disease, Dermatology, Hematology, Rare Disease,
or another appropriate broad therapeutic area.

=== PATTERNS TO LOOK FOR ===
- The trial title often contains the indication
- "in subjects with type 2 diabetes" → T2DM
- Trial acronyms like PIONEER, SUSTAIN, STEP, SELECT often indicate specific conditions
- Outcome studies (e.g., cardiovascular, renal) count as indications

Return ONLY valid JSON:
{{
  "conditions": [
    {{
      "indication": "<disease or condition>",
      "indication_type": "Primary" or "Secondary",
      "therapy_area": "<broad therapeutic area>",
      "rationale": "<why this indication was identified>"
    }}
  ],
  "trial_title": "<trial title>"
}}

Rules:
- Include ALL indications the trial is evaluating — both primary and secondary
- Always extract at least the primary indication from the trial title/details
- Each indication must have indication_type, therapy_area, and rationale
- No explanations outside the JSON
"""
    try:
        resp = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=prompt,
            config=gen_types.GenerateContentConfig(
                temperature=0,
                system_instruction="Return ONLY valid JSON output",
            ),
        )
        text = resp.text.strip()
        text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text).strip()
        data = json.loads(text)
        conditions  = data.get("conditions", [])
        trial_title = data.get("trial_title", "N/A")

        # Normalize to list of dicts with all required fields
        if isinstance(conditions, list):
            normalized = []
            for c in conditions:
                if isinstance(c, dict):
                    normalized.append({
                        "indication":      (c.get("indication") or "").strip(),
                        "indication_type": (c.get("indication_type") or "Secondary").strip(),
                        "therapy_area":    (c.get("therapy_area") or "Other").strip(),
                        "rationale":       (c.get("rationale") or "").strip(),
                    })
                elif isinstance(c, str) and c.strip():
                    normalized.append({
                        "indication":      c.strip(),
                        "indication_type": "Secondary",
                        "therapy_area":    "Other",
                        "rationale":       "",
                    })
            conditions = normalized
        else:
            conditions = [{"indication": str(conditions), "indication_type": "Secondary",
                           "therapy_area": "Other", "rationale": ""}]

        # Deduplicate by indication
        seen, deduped = set(), []
        for c in conditions:
            key = c["indication"].lower()
            if key and key not in seen:
                seen.add(key)
                deduped.append(c)
        conditions = deduped

        print(f"     ✅ {trial_id}: {', '.join(c['indication'] + ' (' + c['indication_type'] + ')' for c in conditions)}")
        with cp_lock:
            checkpoint[trial_id] = {"conditions": conditions, "trial_title": trial_title}
            save_checkpoint(cp_file, checkpoint)

    except Exception as e:
        conditions  = []
        trial_title = str(e)
        print(f"     ❌ {trial_id}: {e}")

    return (trial_id, conditions, trial_title, row)


def _gemini_enrich_trials(rows: list[dict], molecule_name: str,
                          checkpoint_prefix: str, data_source_label: str) -> list[dict]:
    from google import genai
    from google.genai import types as gen_types

    client  = genai.Client(api_key=GEMINI_API_KEY)
    total   = len(rows)
    cp_file = f"checkpoint_{checkpoint_prefix}_{molecule_name.lower().replace(' ', '_')}.json"
    cp      = load_checkpoint(cp_file)
    cp_lock = threading.Lock()

    print(f"  📁 Checkpoint: {len(cp)} completed rows")
    print(f"  🚀 Processing {total} trials with {MAX_WORKERS_TRIALS} parallel workers…\n")

    # ── Parallel enrichment ──────────────────────────────────────────────
    results = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS_TRIALS) as executor:
        futures = {
            executor.submit(
                _enrich_single_trial, row, molecule_name, client, gen_types, cp, cp_file, cp_lock
            ): i
            for i, row in enumerate(rows)
        }
        for future in as_completed(futures):
            idx = futures[future]
            try:
                results.append(future.result())
            except Exception as e:
                row = rows[idx]
                print(f"     ❌ Unexpected: {row.get('trial_id')}: {e}")
                results.append((row.get("trial_id"), [], str(e), row))

    # ── Sort by original order ───────────────────────────────────────────
    trial_order = {row.get("trial_id"): i for i, row in enumerate(rows)}
    results.sort(key=lambda r: trial_order.get(r[0], 0))

    # ── Unnest using LLM classification ──────────────────────────────────
    flat = []
    for trial_id, conditions, trial_title, row in results:
        if not conditions:
            flat.append({
                "molecule_name":   row.get("molecule_name"),
                "company_name":    row.get("company_name"),
                "indication":      "No indication found",
                "rationale":       "",
                "indication_type": "",
                "therapy_area":    "",
                "trial_title":     trial_title,
                "trial_id":        trial_id,
                "phase":           row.get("phase"),
                "source_url":      row.get("source_url"),
                "data_source":     data_source_label,
            })
            continue

        seen = set()
        for c in conditions:
            raw_indication = c.get("indication", "")
            if not raw_indication:
                continue

            # Light normalization
            std = standardize_indication(raw_indication)
            if not std or std.lower() in ("error", "n/a", "no indication found", "none"):
                continue
            if std.lower() in seen:
                continue
            seen.add(std.lower())

            flat.append({
                "molecule_name":   row.get("molecule_name"),
                "company_name":    row.get("company_name"),
                "indication":      std,
                "rationale":       c.get("rationale", ""),
                "indication_type": c.get("indication_type", "Secondary"),
                "therapy_area":    c.get("therapy_area", "Other"),
                "trial_title":     trial_title,
                "trial_id":        trial_id,
                "phase":           row.get("phase"),
                "source_url":      row.get("source_url"),
                "data_source":     data_source_label,
            })

    # ── Compute Ep (row-wise) ────────────────────────────────────────────
    for row in flat:
        row["Ep"] = compute_ep(row.get("phase"))

    # ── Compute Et (drug-level) ──────────────────────────────────────────
    print(f"  🔹 Computing Et (drug-level expansion score)…")
    et_value = compute_et(flat, molecule_name, client, gen_types)
    for row in flat:
        row["Et"] = et_value
    print(f"     Et = {et_value}")

    print(f"\n  ✅ Unnested → {len(flat)} rows (from {total} trials)\n")
    return flat


# ══════════════════════════════════════════════════════════════════════════════
#  MODULE A — label.py  (Clinical Efficacy)
# ══════════════════════════════════════════════════════════════════════════════

def run_label(molecule: str) -> list[dict]:
    print("\n━━━ [1/2] Clinical Efficacy (label.py) ━━━")
    rows = fetch_bq_rows(molecule)
    if not rows:
        print("  ❌ No data found in BigQuery")
        return []
    return _gemini_enrich_trials(rows, molecule, "label", "Clinical Trials")


# ══════════════════════════════════════════════════════════════════════════════
#  MODULE B — drug_indication_researcher.py
# ══════════════════════════════════════════════════════════════════════════════

def _identify_company(molecule: str) -> str:
    from google import genai as _genai
    from google.genai import types as _types
    client = _genai.Client(api_key=GEMINI_API_KEY)
    try:
        resp = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=(
                f"Who is the innovator pharmaceutical company that developed {molecule}? "
                f"Return ONLY JSON: {{\"company\": \"<name>\", \"brand_names\": [\"name1\"]}}"
            ),
            config=_types.GenerateContentConfig(
                temperature=0,
                tools=[_types.Tool(google_search=_types.GoogleSearch())],
                system_instruction="Return ONLY valid JSON. No markdown.",
            ),
        )
        text = re.sub(r"^```(?:json)?\s*|\s*```$", "", resp.text.strip()).strip()
        data = json.loads(text)
        company = data.get("company", "")
        print(f"  🏢 Auto-detected innovator: {company}")
        return company
    except Exception:
        return ""


def _build_innovator_queries(molecule: str, company: str) -> list[dict]:
    return [
        {
            "source_type": "Regulatory Label",
            "prompt": (
                f"Search for official FDA prescribing information and EMA SmPC for {molecule} (by {company}).\n"
                f"For EACH approved indication, return a separate entry.\n\n"
                f"Return ONLY valid JSON:\n"
                f'{{"entries": [\n'
                f'  {{"indication": "...", "brand_name": "...", "source_document": "<exact label title>", '
                f'"phase": "Approved", "source_url": "...", '
                f'"detail": "<approval year, population, dose>", '
                f'"rationale": "<why this indication was identified — cite the specific label section or approval>"}}\n'
                f"]}}\n"
                f"source_document must be the REAL document title. No explanation, only JSON."
            ),
        },
        {
            "source_type": "Investor Presentation",
            "prompt": (
                f"Search for {company}'s latest investor presentations, Capital Markets Day slides, "
                f"R&D day presentations about {molecule}.\n"
                f"For EACH indication mentioned with documented outcomes, return a separate entry.\n\n"
                f"Return ONLY valid JSON:\n"
                f'{{"entries": [\n'
                f'  {{"indication": "...", "brand_name": "...", '
                f'"source_document": "<real presentation title with year>", '
                f'"phase": "<Phase 1/2/3/Filed/Approved/Launched>", '
                f'"source_url": "...", '
                f'"detail": "<specific claim from the presentation>", '
                f'"rationale": "<why this indication was identified — cite the specific evidence>"}}\n'
                f"]}}\n"
                f"Only include indications with observed outcomes data. No explanation, only JSON."
            ),
        },
        {
            "source_type": "Press Release",
            "prompt": (
                f"Search for press releases and earnings call statements from {company} about {molecule}.\n"
                f"For EACH indication mentioned with documented outcomes, return a separate entry.\n\n"
                f"Return ONLY valid JSON:\n"
                f'{{"entries": [\n'
                f'  {{"indication": "...", "brand_name": "...", '
                f'"source_document": "<real press release headline or earnings call date>", '
                f'"phase": "<Phase 1/2/3/Filed/Approved>", '
                f'"source_url": "...", '
                f'"detail": "<key announcement>", '
                f'"rationale": "<why this indication was identified — cite the specific news or data>"}}\n'
                f"]}}\n"
                f"Only include indications backed by reported outcomes. No explanation, only JSON."
            ),
        },
        {
            "source_type": "Pipeline Page",
            "prompt": (
                f"Search for {company}'s official pipeline page listing {molecule}.\n"
                f"For EACH indication on the pipeline, return a separate entry.\n\n"
                f"Return ONLY valid JSON:\n"
                f'{{"entries": [\n'
                f'  {{"indication": "...", "brand_name": "...", '
                f'"source_document": "<e.g. {company} Pipeline — Q1 2025>", '
                f'"phase": "<Preclinical/Phase 1/2/3/Filed/Approved>", '
                f'"source_url": "...", '
                f'"detail": "<status note>", '
                f'"rationale": "<why this indication was identified — cite the pipeline entry>"}}\n'
                f"]}}\n"
                f"Exclude Preclinical entries (no human data). phase MUST be filled. No explanation, only JSON."
            ),
        },
        {
            "source_type": "SEC Filing",
            "prompt": (
                f"Search the SEC EDGAR database and the web for SEC filings by {company} "
                f"(10-K annual reports, 10-Q quarterly reports, 8-K current reports, "
                f"20-F annual reports for foreign private issuers, 6-K reports) "
                f"that disclose clinical or regulatory information about {molecule}.\n\n"
                f"For EACH indication described with actual clinical data or regulatory outcomes "
                f"in these filings, return a separate entry.\n\n"
                f"Return ONLY valid JSON:\n"
                f'{{"entries": [\n'
                f'  {{"indication": "...", "brand_name": "...", '
                f'"source_document": "<exact filing type and period, e.g. {company} 10-K FY2024>", '
                f'"phase": "<Phase 1/2/3/Filed/Approved/Launched>", '
                f'"source_url": "<SEC EDGAR URL>", '
                f'"detail": "<specific disclosed data: trial outcome, milestone, MD&A section>", '
                f'"rationale": "<why this indication was identified — cite the filing section and data>"}}\n'
                f"]}}\n"
                f"Only include indications with actual results reported in the filing — "
                f"not forward-looking statements or boilerplate risk factors. No explanation, only JSON."
            ),
        },
    ]


def _research_single_source(q, molecule, company, client, gen_types):
    """Query a single innovator source. Returns (source_type, rows)."""
    source_type = q["source_type"]
    rows = []
    try:
        resp = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=q["prompt"],
            config=gen_types.GenerateContentConfig(
                temperature=0,
                tools=[gen_types.Tool(google_search=gen_types.GoogleSearch())],
                system_instruction=(
                    "You are a pharmaceutical analyst researching what the "
                    "innovator company publicly says about this drug, including SEC filings. "
                    "Search the web and SEC EDGAR. Return ONLY valid JSON. "
                    "Include indications with documented, verifiable clinical outcomes."
                ),
            ),
        )
        text = re.sub(r"^```(?:json)?\s*|\s*```$", "", resp.text.strip()).strip()
        entries = json.loads(text).get("entries", [])
        if not entries:
            print(f"     ⚠️  {source_type}: No entries")
            return (source_type, [])
        print(f"     ✅ {source_type}: {len(entries)} entries")

        for e in entries:
            raw = e.get("indication", "").strip()
            if not raw:
                continue
            std = standardize_indication(raw)
            if std.lower() in ("error", "n/a", "none", "no indication found"):
                continue
            rationale = (e.get("rationale") or e.get("detail") or "").strip()

            # ── LLM classification for innovator sources ─────────────────
            # For innovator sources, we use a quick LLM call to classify
            ind_type, therapy_area = _classify_single_indication_llm(
                molecule, std, e.get("source_document", ""),
                e.get("phase", ""), client, gen_types
            )

            rows.append({
                "molecule_name":   molecule.title(),
                "company_name":    company,
                "indication":      std,
                "rationale":       rationale,
                "indication_type": ind_type,
                "therapy_area":    therapy_area,
                "trial_title":     e.get("source_document", ""),
                "trial_id":        e.get("brand_name", ""),
                "phase":           e.get("phase", ""),
                "source_url":      e.get("source_url", ""),
                "data_source":     f"Innovator: {source_type}",
            })
    except json.JSONDecodeError:
        print(f"     ⚠️  {source_type}: JSON parse failed")
    except Exception as ex:
        print(f"     ❌ {source_type}: {ex}")

    return (source_type, rows)


def _classify_single_indication_llm(molecule, indication, source_doc, phase,
                                     client, gen_types):
    """Quick LLM call to classify a single indication and assign therapy area."""
    prompt = f"""
You are a pharmaceutical analyst.

Drug: {molecule}
Indication: {indication}
Source: {source_doc}
Phase: {phase}

1. Is this indication "Primary" or "Secondary" for this drug?
   - Primary: the drug's main approved or originally intended indication
   - Secondary: a label expansion beyond the primary indication

2. What is the therapy area?
   Choose from: Metabolic, Cardiovascular, Oncology, Neuroscience, Immunology,
   Respiratory, Nephrology, Hepatology, Ophthalmology, Musculoskeletal,
   Gastroenterology, Infectious Disease, Dermatology, Hematology, Rare Disease, Other

Return ONLY valid JSON:
{{"indication_type": "Primary" or "Secondary", "therapy_area": "<area>"}}
"""
    try:
        resp = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=prompt,
            config=gen_types.GenerateContentConfig(
                temperature=0,
                system_instruction="Return ONLY valid JSON.",
            ),
        )
        text = re.sub(r"^```(?:json)?\s*|\s*```$", "", resp.text.strip()).strip()
        data = json.loads(text)
        return (
            data.get("indication_type", "Secondary"),
            data.get("therapy_area", "Other"),
        )
    except Exception:
        return ("Secondary", "Other")


def _innovator_research(molecule: str, company: str) -> list[dict]:
    from google import genai as _genai
    from google.genai import types as _types

    client   = _genai.Client(api_key=GEMINI_API_KEY)
    queries  = _build_innovator_queries(molecule, company)
    all_rows = []

    print(f"  🚀 Querying {len(queries)} innovator sources in parallel…\n")

    with ThreadPoolExecutor(max_workers=MAX_WORKERS_INNOVATOR) as executor:
        futures = {
            executor.submit(
                _research_single_source, q, molecule, company, client, _types
            ): q["source_type"]
            for q in queries
        }
        for future in as_completed(futures):
            source_type = futures[future]
            try:
                _, rows = future.result()
                all_rows.extend(rows)
            except Exception as ex:
                print(f"     ❌ {source_type}: {ex}")

    seen, unique = set(), []
    for row in all_rows:
        key = (row["indication"], row["trial_title"], row["data_source"])
        if key not in seen:
            seen.add(key)
            unique.append(row)

    # ── Compute Ep for innovator rows ────────────────────────────────────
    for row in unique:
        row["Ep"] = compute_ep(row.get("phase"))

    # ── Compute Et (drug-level) ──────────────────────────────────────────
    if unique:
        print(f"  🔹 Computing Et for innovator rows…")
        et_value = compute_et(unique, molecule, client, _types)
        for row in unique:
            row["Et"] = et_value
        print(f"     Et = {et_value}")

    print(f"\n  📊 Innovator rows: {len(unique)} (deduped from {len(all_rows)})\n")
    return unique


def run_indication_researcher(molecule: str, company: str | None = None) -> list[dict]:
    print("\n━━━ [2/2] Drug Indication Research (Innovator + SEC Sources) ━━━")
    if not company:
        print("  🔎 Identifying innovator company…")
        company = _identify_company(molecule) or ""
    return _innovator_research(molecule, company)


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL WRITER — common format for all sheets
# ══════════════════════════════════════════════════════════════════════════════

def _thin_border():
    return Border(bottom=Side(style="thin", color="DDDDDD"),
                  right=Side(style="thin", color="DDDDDD"))


def _write_standard_sheet(ws, rows: list[dict], molecule: str, sheet_title: str):
    thin     = _thin_border()
    hfill    = PatternFill("solid", start_color="1A1A2E")
    alt_fill = PatternFill("solid", start_color="F4F7FB")
    ncols    = len(HEADERS)

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c           = ws["A1"]
    c.value     = f"{sheet_title}  —  {molecule.title()}"
    c.font      = Font(name="Arial", bold=True, size=14, color="1A1A2E")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    # Subtitle row
    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    c           = ws["A2"]
    c.value     = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}   |   Rows: {len(rows)}"
    c.font      = Font(name="Arial", italic=True, size=9, color="888888")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    # Header row
    for col, h in enumerate(HEADERS, 1):
        c           = ws.cell(row=3, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = hfill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin
    ws.row_dimensions[3].height = 22

    # Data rows
    for idx, row in enumerate(rows, start=4):
        fill = alt_fill if idx % 2 == 0 else None
        for col, key in enumerate(HEADERS, 1):
            c           = ws.cell(row=idx, column=col, value=row.get(key, ""))
            c.font      = Font(name="Arial", size=9)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border    = thin
            if fill:
                c.fill = fill
        ws.row_dimensions[idx].height = 18

    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{3 + len(rows)}"
    ws.freeze_panes = "A4"
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════════════════════════════
#  PER-MOLECULE EXCEL WRITER
# ══════════════════════════════════════════════════════════════════════════════

def write_molecule_excel(molecule: str, label_rows: list[dict],
                         indication_rows: list[dict], out_file: str):
    """Write the 3-sheet Excel for a single molecule."""
    wb = openpyxl.Workbook()

    ws1       = wb.active
    ws1.title = "Clinical Efficacy"
    _write_standard_sheet(ws1, label_rows, molecule, "Clinical Efficacy")

    ws2       = wb.create_sheet("Drug Indication Research")
    _write_standard_sheet(ws2, indication_rows, molecule, "Drug Indication Research")

    combined  = label_rows + indication_rows
    ws3       = wb.create_sheet("All Combined")
    _write_standard_sheet(ws3, combined, molecule, "All Sources Combined")

    wb.save(out_file)
    return combined


# ══════════════════════════════════════════════════════════════════════════════
#  MASTER EXCEL WRITER  (one sheet per molecule + grand summary)
# ══════════════════════════════════════════════════════════════════════════════

def write_master_excel(all_results: dict[str, list[dict]], out_file: str):
    """
    all_results: {molecule_name: [combined rows]}
    Writes one sheet per molecule + a grand 'All Molecules' sheet.
    """
    print(f"\n📊 Writing master file: {out_file}…")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    grand_rows = []

    for molecule, rows in all_results.items():
        # Truncate sheet name to Excel's 31-char limit
        sheet_name = molecule.title()[:31]
        ws = wb.create_sheet(sheet_name)
        _write_standard_sheet(ws, rows, molecule, "All Sources")
        grand_rows.extend(rows)

    ws_all = wb.create_sheet("All Molecules")
    _write_standard_sheet(ws_all, grand_rows, "All Molecules", "Grand Summary")

    wb.save(out_file)
    print(f"✅  Master file saved: {out_file}  ({len(grand_rows)} total rows across {len(all_results)} molecules)")


# ══════════════════════════════════════════════════════════════════════════════
#  SINGLE-MOLECULE PIPELINE
# ══════════════════════════════════════════════════════════════════════════════

def process_molecule(molecule: str, company: str | None = None,
                     output_dir: str = ".") -> tuple[list[dict], list[dict]]:
    """Run both modules for one molecule. Returns (label_rows, indication_rows)."""
    label_rows      = run_label(molecule)
    indication_rows = run_indication_researcher(molecule, company)
    return label_rows, indication_rows


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Unified drug research pipeline → Excel")
    group  = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--molecule", help="Single drug/molecule name, e.g. semaglutide")
    group.add_argument(
        "--drugs",
        nargs="+",
        metavar="DRUG",
        help="One or more drug names, e.g. --drugs Semaglutide Liraglutide Tirzepatide",
    )
    group.add_argument("--all", action="store_true",
                       help="Run pipeline for ALL distinct molecules in BigQuery")
    parser.add_argument("--company", default=None,
                        help="Innovator company (only used with --molecule)")
    parser.add_argument("--output-dir", default=".",
                        help="Directory to write output files (default: current dir)")
    parser.add_argument("--skip-existing", action="store_true",
                        help="(--all / --drugs mode) skip molecules whose Excel file already exists")
    parser.add_argument("--master-file", default=None,
                        help="(--all mode) filename for the master combined Excel "
                             "(default: all_molecules_research_<timestamp>.xlsx)")
    args = parser.parse_args()

    if not GEMINI_API_KEY:
        sys.exit("❌  GEMINI_API_KEY not set in .env")

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # ── SINGLE-MOLECULE MODE ──────────────────────────────────────────────
    if args.molecule:
        molecule = args.molecule.strip()
        slug     = molecule.lower().replace(" ", "_")
        out_file = output_dir / f"{slug}_research.xlsx"

        print(f"\n{'━'*62}")
        print(f"  Molecule : {molecule.title()}")
        print(f"  Output   : {out_file}")
        print(f"{'━'*62}")

        label_rows, indication_rows = process_molecule(molecule, args.company)
        combined = write_molecule_excel(molecule, label_rows, indication_rows, str(out_file))

        print(f"\n✅  Done!")
        print(f"📄  File   : {out_file}")
        print(f"📊  Sheets : Clinical Efficacy ({len(label_rows)}) | "
              f"Drug Indication Research ({len(indication_rows)}) | "
              f"All Combined ({len(combined)})")
        return

    # ── MULTI-DRUG MODE  (--drugs) ────────────────────────────────────────
    if args.drugs:
        molecules = [m.strip() for m in args.drugs if m.strip()]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        slug      = "_".join(m.lower().replace(" ", "_") for m in molecules[:3])
        if len(molecules) > 3:
            slug += f"_and_{len(molecules) - 3}_more"
        master_file = args.master_file or f"{slug}_research_{timestamp}.xlsx"
        master_path = output_dir / master_file

        print(f"\n{'━'*62}")
        print(f"  Mode     : MULTI-DRUG  ({len(molecules)} drugs)")
        print(f"  Drugs    : {', '.join(m.title() for m in molecules)}")
        print(f"  Output   : {output_dir}")
        print(f"  Master   : {master_path}")
        print(f"{'━'*62}\n")

        all_results: dict[str, list[dict]] = {}
        failed:      list[str]             = []

        total = len(molecules)
        for i, molecule in enumerate(molecules, 1):
            slug     = molecule.lower().replace(" ", "_")
            out_file = output_dir / f"{slug}_research.xlsx"

            print(f"\n{'═'*62}")
            print(f"  [{i}/{total}] {molecule.title()}")
            print(f"{'═'*62}")

            if args.skip_existing and out_file.exists():
                print(f"  ⏭️  Skipping — file already exists: {out_file}")
                all_results[molecule] = []
                continue

            try:
                label_rows, indication_rows = process_molecule(molecule)
                combined = write_molecule_excel(molecule, label_rows, indication_rows, str(out_file))
                all_results[molecule] = combined

                print(f"\n  ✅ {molecule.title()} done → {out_file}")
                print(f"     Clinical Efficacy: {len(label_rows)} | "
                      f"Indication Research: {len(indication_rows)} | "
                      f"Combined: {len(combined)}")
            except Exception as e:
                print(f"\n  ❌ FAILED: {molecule} — {e}")
                failed.append(molecule)
                all_results[molecule] = []

        write_master_excel(all_results, str(master_path))

        print(f"\n{'━'*62}")
        print(f"  Pipeline complete!")
        print(f"  Drugs processed : {total}")
        print(f"  Succeeded       : {total - len(failed)}")
        print(f"  Failed          : {len(failed)}")
        if failed:
            print(f"  Failed drugs    : {', '.join(failed)}")
        print(f"  Master file     : {master_path}")
        print(f"  Per-drug files  : {output_dir}/<drug>_research.xlsx")
        print(f"{'━'*62}\n")
        return

    # ── ALL-MOLECULES MODE ────────────────────────────────────────────────
    print(f"\n{'━'*62}")
    print(f"  Mode     : ALL MOLECULES")
    print(f"  Output   : {output_dir}")
    print(f"{'━'*62}\n")

    molecules = fetch_all_molecules()
    if not molecules:
        sys.exit("❌  No molecules found in BigQuery.")

    timestamp    = datetime.now().strftime("%Y%m%d_%H%M%S")
    master_file  = args.master_file or f"all_molecules_research_{timestamp}.xlsx"
    master_path  = output_dir / master_file

    all_results: dict[str, list[dict]] = {}
    failed:      list[str]             = []

    total = len(molecules)
    for i, molecule in enumerate(molecules, 1):
        slug     = molecule.lower().replace(" ", "_")
        out_file = output_dir / f"{slug}_research.xlsx"

        print(f"\n{'═'*62}")
        print(f"  [{i}/{total}] {molecule.title()}")
        print(f"{'═'*62}")

        # Skip if already done
        if args.skip_existing and out_file.exists():
            print(f"  ⏭️  Skipping — file already exists: {out_file}")
            # Still try to load existing combined rows for the master file
            try:
                wb_existing = openpyxl.load_workbook(str(out_file), read_only=True, data_only=True)
                ws_comb = wb_existing["All Combined"]
                rows_existing = []
                headers_row = None
                for row_cells in ws_comb.iter_rows(min_row=3, values_only=True):
                    if headers_row is None:
                        headers_row = list(row_cells)
                        continue
                    if any(c for c in row_cells):
                        rows_existing.append(dict(zip(HEADERS, row_cells)))
                all_results[molecule] = rows_existing
                wb_existing.close()
            except Exception as e:
                print(f"  ⚠️  Could not read existing file for master: {e}")
                all_results[molecule] = []
            continue

        try:
            label_rows, indication_rows = process_molecule(molecule)
            combined = write_molecule_excel(molecule, label_rows, indication_rows, str(out_file))
            all_results[molecule] = combined

            print(f"\n  ✅ {molecule.title()} done → {out_file}")
            print(f"     Clinical Efficacy: {len(label_rows)} | "
                  f"Indication Research: {len(indication_rows)} | "
                  f"Combined: {len(combined)}")

        except Exception as e:
            print(f"\n  ❌ FAILED: {molecule} — {e}")
            failed.append(molecule)
            all_results[molecule] = []

    # Write master file
    write_master_excel(all_results, str(master_path))

    # Summary
    print(f"\n{'━'*62}")
    print(f"  Pipeline complete!")
    print(f"  Molecules processed : {total}")
    print(f"  Succeeded           : {total - len(failed)}")
    print(f"  Failed              : {len(failed)}")
    if failed:
        print(f"  Failed molecules    : {', '.join(failed)}")
    print(f"  Master file         : {master_path}")
    print(f"  Per-molecule files  : {output_dir}/<molecule>_research.xlsx")
    print(f"{'━'*62}\n")


if __name__ == "__main__":
    main()
