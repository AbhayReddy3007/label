"""Writes the single-drug label-expansion workbook.

Takes the row lists produced by `research_modules.py` (Clinical Efficacy
rows, Drug Indication Research rows, OpenTargets rows) and writes them to
one .xlsx with five sheets: Clinical Efficacy, Drug Indication Research,
OpenTargets Indications, All Combined, and Summary. This module knows
nothing about BigQuery, GCS, or Gemini — it only formats rows that are
already plain dicts.
"""

from __future__ import annotations

import logging
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import medical_potential.config as config

logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════
#  COMMON OUTPUT FORMAT
# ══════════════════════════════════════════════════════════════════════════
HEADERS = [
    "molecule_name", "company_name", "indication", "rationale",
    "indication_type", "therapy_area", "TA-I", "trial_title", "trial_id",
    "trial_size", "phase", "source_url", "data_source", "Ep", "Et", "moa",
    "opentargets_score",
]
COL_WIDTHS = [18, 22, 28, 40, 16, 18, 36, 50, 18, 12, 10, 40, 16, 8, 8, 28, 18]

# Headers for the dedicated OpenTargets sheet
OT_HEADERS = [
    "molecule_name", "company_name", "indication", "rationale",
    "therapy_area", "moa", "opentargets_score",
    "trial_title", "trial_id", "source_url", "data_source",
]
OT_COL_WIDTHS = [18, 22, 28, 40, 24, 28, 18, 50, 18, 40, 16]


def _thin_border() -> Border:
    return Border(bottom=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"))


def _write_standard_sheet(ws, rows: list[dict], molecule: str, sheet_title: str,
                          headers: list[str] | None = None,
                          col_widths: list[int] | None = None) -> None:
    headers = headers or HEADERS
    col_widths = col_widths or COL_WIDTHS
    thin = _thin_border()
    hfill = PatternFill("solid", start_color="1A1A2E")
    alt_fill = PatternFill("solid", start_color="F4F7FB")
    ncols = len(headers)

    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value = f"{sheet_title}  —  {molecule.title()}"
    c.font = Font(name="Arial", bold=True, size=14, color="1A1A2E")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    c = ws["A2"]
    c.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}   |   Rows: {len(rows)}"
    c.font = Font(name="Arial", italic=True, size=9, color="888888")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin
    ws.row_dimensions[3].height = 22

    for idx, row in enumerate(rows, start=4):
        fill = alt_fill if idx % 2 == 0 else None
        for col, key in enumerate(headers, 1):
            val = row.get(key, "")
            # Format opentargets_score as number if present
            if key == "opentargets_score" and isinstance(val, (int, float)) and val:
                c = ws.cell(row=idx, column=col, value=round(val, 4))
                c.number_format = '0.0000'
            else:
                c = ws.cell(row=idx, column=col, value=val)
            c.font = Font(name="Arial", size=9)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border = thin
            if fill:
                c.fill = fill
        ws.row_dimensions[idx].height = 18

    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{3 + len(rows)}"
    ws.freeze_panes = "A4"
    for i, w in enumerate(col_widths, 1):
        if i <= len(col_widths):
            ws.column_dimensions[get_column_letter(i)].width = w


def _write_summary_sheet(wb, rows: list[dict], title_label: str = "Summary",
                         moa_list: list[str] | None = None) -> None:
    """Add a 'Summary' sheet: Drug Name | No. of Indications | No. of Therapy Areas | MOA."""
    ws = wb.create_sheet("Summary")
    thin = _thin_border()
    hfill = PatternFill("solid", start_color="1A1A2E")
    alt = PatternFill("solid", start_color="F4F7FB")

    drug_data = defaultdict(lambda: {"display_name": "", "indications": set(), "therapy_areas": set()})
    for r in rows:
        name = (r.get("molecule_name") or "").strip()
        ind = r.get("indication", "")
        ta = r.get("therapy_area", "")
        if not name or ind == "No indication found":
            continue
        entry = drug_data[name.lower()]
        if not entry["display_name"]:
            entry["display_name"] = name
        entry["indications"].add(ind)
        if ta and ta.lower() not in ("other", ""):
            entry["therapy_areas"].add(ta)

    moa_str = "; ".join(moa_list) if moa_list else ""

    summary_headers = ["Drug Name", "No. of Indications", "No. of Therapy Areas", "Mechanism of Action"]
    summary_widths = [28, 20, 22, 50]

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = f"Summary  —  {title_label}"
    c.font = Font(name="Arial", bold=True, size=14, color="1A1A2E")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    for col, h in enumerate(summary_headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin
    ws.row_dimensions[2].height = 22

    for idx, (_, info) in enumerate(sorted(drug_data.items()), start=3):
        fill = alt if idx % 2 == 1 else None
        vals = [info["display_name"], len(info["indications"]), len(info["therapy_areas"]), moa_str]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=idx, column=col, value=val)
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(
                horizontal="center" if col in (2, 3) else "left",
                vertical="center",
                wrap_text=True,
            )
            c.border = thin
            if fill:
                c.fill = fill
        ws.row_dimensions[idx].height = 20

    for i, w in enumerate(summary_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_excel(
    molecule: str,
    clinical_rows: list[dict],
    indication_rows: list[dict],
    opentargets_rows: list[dict] | None = None,
    moa_list: list[str] | None = None,
) -> Path:
    """Write the single-drug workbook: Clinical Efficacy, Drug Indication
    Research, OpenTargets Indications, All Combined, and Summary sheets.

    Writes directly into config.OUTPUT_DIR (created once, reused across
    runs) as `<drug>_label_expansion.xlsx` — no per-run output folder.
    """
    opentargets_rows = opentargets_rows or []
    moa_list = moa_list or []

    # Ensure moa and opentargets_score keys exist on all rows
    for row in clinical_rows + indication_rows:
        row.setdefault("moa", "")
        row.setdefault("opentargets_score", "")
        row.setdefault("trial_size", "")

    output_dir = Path(config.OUTPUT_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)
    slug = molecule.lower().replace(" ", "_")
    out_file = output_dir / f"{slug}_label_expansion.xlsx"

    wb = openpyxl.Workbook()

    # ── Sheet 1: Clinical Efficacy ────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Clinical Efficacy"
    _write_standard_sheet(ws1, clinical_rows, molecule, "Clinical Efficacy")

    # ── Sheet 2: Drug Indication Research ─────────────────────────────
    ws2 = wb.create_sheet("Drug Indication Research")
    _write_standard_sheet(ws2, indication_rows, molecule, "Drug Indication Research")

    # ── Sheet 3: OpenTargets Indications ──────────────────────────────
    ws3 = wb.create_sheet("OpenTargets Indications")
    _write_standard_sheet(ws3, opentargets_rows, molecule, "OpenTargets Indications",
                          headers=OT_HEADERS, col_widths=OT_COL_WIDTHS)

    # ── Sheet 4: All Combined ─────────────────────────────────────────
    combined = clinical_rows + indication_rows + opentargets_rows

    # Et in "All Combined" = highest Et across both modules
    et_values = [r.get("Et") for r in combined if r.get("Et")]
    max_et = ""
    if et_values:
        try:
            max_et = str(max(int(v) for v in et_values if str(v).isdigit()))
        except ValueError:
            max_et = et_values[0]
    for row in combined:
        row["Et"] = max_et

    ws4 = wb.create_sheet("All Combined")
    _write_standard_sheet(ws4, combined, molecule, "All Sources Combined")

    # ── Sheet 5: Summary ──────────────────────────────────────────────
    _write_summary_sheet(wb, combined, molecule.title(), moa_list=moa_list)

    wb.save(str(out_file))
    logger.info("Excel workbook saved → %s", out_file)
    return out_file
